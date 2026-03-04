"""
桑名市FAQ不足質問ピックアップツール（ハイブリッド版）

処理フロー:
  ① 代表質問一覧・総務省FAQ・桑名市FAQを読み込み
  ② Gemini Embedding APIで全質問をベクトル化
  ③ Embedding類似度で機械的にフィルタリング
     - 総務省FAQとの最大類似度が高い → 国の制度FAQでカバー済み → 除外候補
     - 桑名市FAQとの最大類似度が高い → 桑名市で既にカバー済み → 除外候補
  ④ グレーゾーンの質問をLLM（Gemini）で精密判定
  ⑤ 結果をExcelで出力

使い方:
  export GEMINI_API_KEY="your-api-key"
  python faq_gap_finder.py \
    --representative 代表質問一覧50件.xlsx \
    --soumu 総務省FAQ.xlsx \
    --kuwana 桑名市FAQ.xlsx \
    --output 桑名市_不足質問一覧.xlsx

必要パッケージ:
  pip install pandas openpyxl numpy scikit-learn aiohttp
"""

from dotenv import load_dotenv
load_dotenv()

import os
import re
import sys
import asyncio
import argparse
import numpy as np
import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity

# ============================================================
# 設定
# ============================================================
from pathlib import Path
_BASE_DIR = Path(__file__).parent

GEMINI_EMBED_MODEL = "gemini-embedding-001"
GEMINI_LLM_MODEL = "gemini-3.1-pro-preview"

# デフォルトファイルパス
DEFAULT_SOUMU_FILE = str(_BASE_DIR / "総務省FAQ.xlsx")
DEFAULT_KUWANA_FILE = str(_BASE_DIR / "桑名市FAQ.xlsx")

# Embedding API パラメータ
DEFAULT_BATCH_SIZE = 50
DEFAULT_MAX_CONCURRENT = 5
DEFAULT_RATE_LIMIT_DELAY = 0.5

# 類似度閾値
DEFAULT_SOUMU_HIGH = 0.90     # 総務省FAQとの類似度がこれ以上 → 自動除外
DEFAULT_SOUMU_LOW = 0.60      # これ以下 → 総務省には無い（自動通過）
DEFAULT_KUWANA_HIGH = 0.90    # 桑名市FAQとの類似度がこれ以上 → 自動除外（カバー済み）
DEFAULT_KUWANA_LOW = 0.60     # これ以下 → 桑名市に無い（不足確定）


# ============================================================
# ① データ読み込み
# ============================================================
def load_representative(filepath: str) -> pd.DataFrame:
    """代表質問一覧を読み込む"""
    xl = pd.ExcelFile(filepath)
    dfs = []
    for sheet in xl.sheet_names:
        if sheet == "サマリー":
            continue
        df = pd.read_excel(filepath, sheet_name=sheet)
        if "質問" in df.columns:
            dfs.append(df)
    result = pd.concat(dfs, ignore_index=True)
    print(f"代表質問: {len(result)}件")
    return result


def load_soumu(filepath: str) -> pd.DataFrame:
    """総務省FAQを読み込む"""
    xl = pd.ExcelFile(filepath)
    dfs = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet)
        if "質問" in df.columns:
            df["カテゴリ"] = sheet.replace(" FAQ", "")
            dfs.append(df)
    result = pd.concat(dfs, ignore_index=True)
    print(f"総務省FAQ: {len(result)}件")
    return result


def load_kuwana(filepath: str) -> pd.DataFrame:
    """桑名市FAQを読み込む"""
    xl = pd.ExcelFile(filepath)
    dfs = []
    for sheet in xl.sheet_names:
        if sheet == "サマリー":
            continue
        df = pd.read_excel(filepath, sheet_name=sheet)
        if "質問" in df.columns:
            df["カテゴリ"] = sheet
            dfs.append(df)
    result = pd.concat(dfs, ignore_index=True)
    print(f"桑名市FAQ: {len(result)}件")
    return result


# ============================================================
# Embedding用テキスト構築（コンテキスト付与で短文精度を改善）
# ============================================================
DEFAULT_MAX_ANSWER_LEN = 200


def build_embed_text(row, max_answer_len=DEFAULT_MAX_ANSWER_LEN) -> str:
    """質問文にカテゴリ・サブカテゴリ・回答冒頭を付与してEmbedding精度を向上させる。

    短い質問文だけではEmbeddingの意味解像度が低いため、
    周辺情報を結合して情報量を補強する。

    例:
      入力: 質問="ごみ袋について", カテゴリ="環境・ごみ", 回答="指定ごみ袋は..."
      出力: "【環境・ごみ】ごみ袋について。指定ごみ袋は..."
    """
    parts = []

    # カテゴリ・サブカテゴリをラベルとして付与
    cat = str(row.get("統一カテゴリ", row.get("カテゴリ", ""))).strip()
    sub = str(row.get("サブカテゴリ", "")).strip()
    if cat and cat != "nan":
        label = f"【{cat}"
        if sub and sub != "nan":
            label += f"／{sub}"
        label += "】"
        parts.append(label)

    # 質問文（必須）
    q = str(row.get("質問", "")).strip()
    if q and q != "nan":
        parts.append(q)

    # 回答文の冒頭を付与（情報量の補強）
    ans = str(row.get("回答", row.get("回答（参考）", ""))).strip()
    if ans and ans != "nan":
        # HTMLタグ・URL・改行を除去してクリーンなテキストにする
        import re as _re
        ans = _re.sub(r'<[^>]+>', '', ans)
        ans = _re.sub(r'https?://\S+', '', ans)
        ans = _re.sub(r'[\n\r\s]+', ' ', ans).strip()
        if len(ans) > max_answer_len:
            ans = ans[:max_answer_len] + "…"
        parts.append(ans)

    return " ".join(parts)


def build_embed_texts_from_df(df: pd.DataFrame,
                              max_answer_len=DEFAULT_MAX_ANSWER_LEN) -> list[str]:
    """DataFrameの全行からEmbedding用テキストを生成する"""
    return [build_embed_text(row, max_answer_len) for _, row in df.iterrows()]


# ============================================================
# ② Gemini Embedding API（非同期並列処理）
# ============================================================
def embed_with_gemini(texts: list[str], api_key: str,
                      batch_size: int = DEFAULT_BATCH_SIZE,
                      max_concurrent: int = DEFAULT_MAX_CONCURRENT,
                      rate_limit_delay: float = DEFAULT_RATE_LIMIT_DELAY) -> np.ndarray:
    try:
        import aiohttp
    except ImportError:
        print("ERROR: aiohttp が必要です。  pip install aiohttp")
        sys.exit(1)

    async def _run():
        batches = [texts[i:i+batch_size] for i in range(0, len(texts), batch_size)]
        semaphore = asyncio.Semaphore(max_concurrent)
        print(f"  Embedding: {len(texts)}件 → {len(batches)}バッチ "
              f"(バッチサイズ{batch_size}, 並列数{max_concurrent})")

        async def process_batch(batch_texts, session, batch_id):
            url = (f"https://generativelanguage.googleapis.com/v1beta/"
                   f"models/{GEMINI_EMBED_MODEL}:batchEmbedContents?key={api_key}")
            body = {"requests": [
                {"model": f"models/{GEMINI_EMBED_MODEL}",
                 "content": {"parts": [{"text": t}]},
                 "taskType": "SEMANTIC_SIMILARITY"}
                for t in batch_texts
            ]}

            for attempt in range(3):
                async with semaphore:
                    try:
                        async with session.post(url, json=body) as resp:
                            if resp.status == 429:
                                wait = float(resp.headers.get("Retry-After", 5))
                                print(f"  [RATE LIMIT] バッチ{batch_id}: "
                                      f"{wait}秒待機 (リトライ{attempt+1}/3)")
                                await asyncio.sleep(wait)
                                continue
                            elif resp.status != 200:
                                error = await resp.text()
                                print(f"  [ERROR] バッチ{batch_id}: "
                                      f"status={resp.status}, {error[:200]}")
                                return [None] * len(batch_texts)
                            data = await resp.json()
                            return [e.get("values")
                                    for e in data.get("embeddings", [])]
                    except Exception as e:
                        print(f"  [ERROR] バッチ{batch_id}: {e}")
                        if attempt < 2:
                            await asyncio.sleep(2)
                    finally:
                        await asyncio.sleep(rate_limit_delay)
            return [None] * len(batch_texts)

        async with aiohttp.ClientSession() as session:
            tasks = [process_batch(b, session, i)
                     for i, b in enumerate(batches)]
            results = await asyncio.gather(*tasks)

        all_embs, failed, dim = [], 0, None
        for batch_result in results:
            for emb in batch_result:
                if emb is None:
                    failed += 1
                    all_embs.append(None)
                else:
                    if dim is None:
                        dim = len(emb)
                    all_embs.append(np.array(emb))

        if dim is None:
            raise RuntimeError("全Embeddingに失敗しました。APIキーを確認してください。")

        for i in range(len(all_embs)):
            if all_embs[i] is None:
                all_embs[i] = np.zeros(dim)

        if failed:
            print(f"  ⚠ {failed}件のEmbeddingに失敗（ゼロベクトルで代替）")

        print(f"  Embedding完了: {len(all_embs)}件, 次元数: {dim}")
        return np.array(all_embs)

    return asyncio.run(_run())


# ============================================================
# ③ Embedding類似度による機械的フィルタリング
# ============================================================
def compute_max_similarity(emb_source: np.ndarray,
                           emb_target: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    """sourceの各質問に対する、targetとの最大コサイン類似度と最類似インデックスを返す"""
    sim = cosine_similarity(emb_source, emb_target)  # (n_source, n_target)
    max_sim = sim.max(axis=1)       # (n_source,)
    max_idx = sim.argmax(axis=1)    # (n_source,)
    return max_sim, max_idx


# ============================================================
# ④ LLMによるグレーゾーン判定
# ============================================================
def llm_judge_batch(questions: list[dict], api_key: str) -> list[dict]:
    """グレーゾーンの質問をLLMで精密判定する

    各questionは:
      {
        "question": "代表質問のテキスト",
        "zone": "soumu_gray" | "kuwana_gray",
        "similar_question": "最も類似している比較先の質問",
        "similarity": 0.78,
        "category": "カテゴリ名",
      }

    Returns:
      各質問に "llm_verdict" ("除外" or "採用") と "llm_reason" を追加
    """
    from google import genai
    from google.genai import types

    client = genai.Client(api_key=api_key)
    results = []

    # バッチで処理（10件ずつ）
    for i in range(0, len(questions), 10):
        batch = questions[i:i+10]

        items_text = ""
        for j, q in enumerate(batch):
            zone_label = "総務省FAQ" if q["zone"] == "soumu_gray" else "桑名市FAQ"
            items_text += f"""
--- 質問{j+1} ---
代表質問: 「{q['question']}」
比較先（{zone_label}）: 「{q['similar_question']}」
類似度: {q['similarity']:.3f}
"""

        prompt = f"""あなたは自治体FAQの専門家です。
以下の質問ペアを見て、「代表質問」が「比較先の質問」で実質的にカバーされているかを判定してください。

判定基準:
- 「カバー済み」: 比較先の質問で回答すれば、代表質問の回答にもなる（同じ趣旨・同じ情報）
- 「別の質問」: 表面的に似ているが、実際には異なる情報が必要（制度の違い、対象の違い等）

{items_text}

【出力形式】
各質問について、以下の形式で1行ずつ出力してください。他の説明は不要です。
質問1: カバー済み|理由を20文字以内で
質問2: 別の質問|理由を20文字以内で
..."""

        try:
            response = client.models.generate_content(
                model=GEMINI_LLM_MODEL,
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=500,
                ),
            )

            lines = response.text.strip().split("\n")
            for j, q in enumerate(batch):
                verdict = "採用"  # デフォルト（判定不能時は安全側＝採用）
                reason = "LLM判定不能"

                for line in lines:
                    if f"質問{j+1}" in line:
                        if "カバー済み" in line:
                            verdict = "除外"
                            reason = line.split("|")[-1].strip() if "|" in line else "カバー済み"
                        else:
                            verdict = "採用"
                            reason = line.split("|")[-1].strip() if "|" in line else "別の質問"
                        break

                q["llm_verdict"] = verdict
                q["llm_reason"] = reason
                results.append(q)

            print(f"  LLM判定: {i+1}〜{i+len(batch)}件目 完了")

        except Exception as e:
            print(f"  [ERROR] LLM判定失敗: {e}")
            for q in batch:
                q["llm_verdict"] = "採用"
                q["llm_reason"] = f"LLMエラー: {str(e)[:30]}"
                results.append(q)

    return results


# ============================================================
# ⑤ Excel出力
# ============================================================
def save_results(df_result: pd.DataFrame, stats: dict, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="2E75B6")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side("thin"), right=Side("thin"),
                top=Side("thin"), bottom=Side("thin"))

    # 色分け用
    fill_shortage = PatternFill("solid", fgColor="FCE4EC")   # 不足（赤系）
    fill_covered = PatternFill("solid", fgColor="E8F5E9")    # カバー済み（緑系）
    fill_gray = PatternFill("solid", fgColor="FFF8E1")       # グレーゾーン（黄系）

    # --- サマリーシート ---
    ws = wb.create_sheet("サマリー", 0)
    summary_headers = ["項目", "件数"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    summary_data = [
        ("代表質問（入力）", stats["total_representative"]),
        ("総務省FAQでカバー済み（自動除外）", stats["soumu_auto_exclude"]),
        ("総務省FAQグレーゾーン→LLM除外", stats["soumu_llm_exclude"]),
        ("桑名市FAQでカバー済み（自動除外）", stats["kuwana_auto_exclude"]),
        ("桑名市FAQグレーゾーン→LLM除外", stats["kuwana_llm_exclude"]),
        ("桑名市で不足している質問（最終出力）", stats["final_shortage"]),
    ]
    for ri, (label, val) in enumerate(summary_data, 2):
        ws.cell(row=ri, column=1, value=label).font = cf
        ws.cell(row=ri, column=1, value=label).border = tb
        ws.cell(row=ri, column=2, value=val).font = cf
        ws.cell(row=ri, column=2, value=val).border = tb

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12

    # --- 不足質問シート（最終結果）---
    df_shortage = df_result[df_result["最終判定"] == "不足"].copy()
    ws = wb.create_sheet("不足質問一覧")
    headers = ["No.", "カテゴリ", "サブカテゴリ", "質問",
               "総務省FAQ最類似", "総務省類似度",
               "桑名市FAQ最類似", "桑名市類似度",
               "判定根拠", "出典自治体"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for ri, (_, row) in enumerate(df_shortage.reset_index(drop=True).iterrows(), 2):
        vals = [
            ri - 1,
            row.get("統一カテゴリ", ""),
            row.get("サブカテゴリ", ""),
            row["質問"],
            row.get("総務省_最類似質問", ""),
            f'{row.get("総務省_類似度", 0):.3f}',
            row.get("桑名市_最類似質問", ""),
            f'{row.get("桑名市_類似度", 0):.3f}',
            row.get("判定根拠", ""),
            row.get("出典自治体", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.alignment, c.border = cf, ca, tb
            c.fill = fill_shortage

    for ci, w in enumerate([5, 14, 14, 45, 45, 10, 45, 10, 20, 10], 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    if len(df_shortage) > 0:
        ws.auto_filter.ref = f"A1:J{len(df_shortage)+1}"

    # --- 全件詳細シート ---
    ws = wb.create_sheet("全件詳細")
    detail_headers = ["No.", "カテゴリ", "質問",
                      "総務省_最類似質問", "総務省_類似度",
                      "桑名市_最類似質問", "桑名市_類似度",
                      "判定段階", "最終判定", "判定根拠", "出典自治体"]
    for ci, h in enumerate(detail_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for ri, (_, row) in enumerate(df_result.iterrows(), 2):
        vals = [
            ri - 1,
            row.get("統一カテゴリ", ""),
            row["質問"],
            row.get("総務省_最類似質問", ""),
            f'{row.get("総務省_類似度", 0):.3f}',
            row.get("桑名市_最類似質問", ""),
            f'{row.get("桑名市_類似度", 0):.3f}',
            row.get("判定段階", ""),
            row.get("最終判定", ""),
            row.get("判定根拠", ""),
            row.get("出典自治体", ""),
        ]
        fill = fill_shortage if row.get("最終判定") == "不足" else fill_covered
        if "グレーゾーン" in str(row.get("判定段階", "")):
            fill = fill_gray if row.get("最終判定") == "不足" else fill_covered

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.alignment, c.border = cf, ca, tb
            c.fill = fill

    for ci, w in enumerate([5, 16, 45, 45, 10, 45, 10, 16, 8, 20, 10], 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
    if len(df_result) > 0:
        ws.auto_filter.ref = f"A1:K{len(df_result)+1}"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n出力完了: {output_path}")


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="桑名市FAQ不足質問ピックアップツール")
    parser.add_argument("--representative", "-r", required=True,
                        help="代表質問一覧Excelファイル")
    parser.add_argument("--soumu", "-s", default=DEFAULT_SOUMU_FILE,
                        help=f"総務省FAQ Excelファイル (default: {Path(DEFAULT_SOUMU_FILE).name})")
    parser.add_argument("--kuwana", "-k", default=DEFAULT_KUWANA_FILE,
                        help=f"桑名市FAQ Excelファイル (default: {Path(DEFAULT_KUWANA_FILE).name})")
    parser.add_argument("--output", "-o", default="./桑名市_不足質問一覧.xlsx",
                        help="出力Excelファイルパス")
    parser.add_argument("--api-key",
                        help="Gemini APIキー（環境変数 GEMINI_API_KEY でも可）")
    parser.add_argument("--soumu-high", type=float, default=DEFAULT_SOUMU_HIGH,
                        help=f"総務省FAQ自動除外閾値 (default: {DEFAULT_SOUMU_HIGH})")
    parser.add_argument("--soumu-low", type=float, default=DEFAULT_SOUMU_LOW,
                        help=f"総務省FAQ自動通過閾値 (default: {DEFAULT_SOUMU_LOW})")
    parser.add_argument("--kuwana-high", type=float, default=DEFAULT_KUWANA_HIGH,
                        help=f"桑名市FAQ自動除外閾値 (default: {DEFAULT_KUWANA_HIGH})")
    parser.add_argument("--kuwana-low", type=float, default=DEFAULT_KUWANA_LOW,
                        help=f"桑名市FAQ自動通過閾値 (default: {DEFAULT_KUWANA_LOW})")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--max-concurrent", type=int, default=DEFAULT_MAX_CONCURRENT)
    parser.add_argument("--max-answer-len", type=int, default=DEFAULT_MAX_ANSWER_LEN,
                        help=f"Embedding用に結合する回答文の最大文字数 (default: {DEFAULT_MAX_ANSWER_LEN})")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: Gemini APIキーが必要です。")
        print("  --api-key YOUR_KEY または export GEMINI_API_KEY='YOUR_KEY'")
        sys.exit(1)

    print("=" * 60)
    print("桑名市FAQ 不足質問ピックアップツール")
    print(f"  総務省FAQ閾値: 自動除外={args.soumu_high}, "
          f"グレーゾーン={args.soumu_low}〜{args.soumu_high}, "
          f"自動通過=〜{args.soumu_low}")
    print(f"  桑名市FAQ閾値: 自動除外={args.kuwana_high}, "
          f"グレーゾーン={args.kuwana_low}〜{args.kuwana_high}, "
          f"自動通過=〜{args.kuwana_low}")
    print("=" * 60)

    # ① 読み込み
    print("\n【STEP 1】データ読み込み")
    df_rep = load_representative(args.representative)
    df_soumu = load_soumu(args.soumu)
    df_kuwana = load_kuwana(args.kuwana)

    # ② Embedding
    print("\n【STEP 2】Embedding（コンテキスト付与 + 一括ベクトル化）")
    print("  質問文にカテゴリ・サブカテゴリ・回答冒頭を結合してEmbedding精度を向上")

    rep_embed_texts = build_embed_texts_from_df(df_rep)
    soumu_embed_texts = build_embed_texts_from_df(df_soumu)
    kuwana_embed_texts = build_embed_texts_from_df(df_kuwana)

    # Embedding用テキストのサンプル表示
    print(f"\n  [サンプル] 代表質問:")
    print(f"    元の質問: {df_rep.iloc[0]['質問']}")
    print(f"    Embedding用: {rep_embed_texts[0][:120]}...")
    print(f"  [サンプル] 総務省FAQ:")
    print(f"    元の質問: {df_soumu.iloc[0]['質問']}")
    print(f"    Embedding用: {soumu_embed_texts[0][:120]}...")

    all_texts = rep_embed_texts + soumu_embed_texts + kuwana_embed_texts
    print(f"\n  合計: {len(all_texts)}件 "
          f"(代表{len(rep_embed_texts)} + 総務省{len(soumu_embed_texts)} "
          f"+ 桑名市{len(kuwana_embed_texts)})")

    all_emb = embed_with_gemini(
        all_texts, api_key,
        batch_size=args.batch_size,
        max_concurrent=args.max_concurrent,
    )

    # Embeddingを分割
    n_rep = len(rep_embed_texts)
    n_soumu = len(soumu_embed_texts)
    emb_rep = all_emb[:n_rep]
    emb_soumu = all_emb[n_rep:n_rep + n_soumu]
    emb_kuwana = all_emb[n_rep + n_soumu:]

    # ③ 類似度計算・機械的フィルタリング
    print("\n【STEP 3】類似度計算・機械的フィルタリング")

    # 総務省FAQとの照合
    soumu_sim, soumu_idx = compute_max_similarity(emb_rep, emb_soumu)
    # 桑名市FAQとの照合
    kuwana_sim, kuwana_idx = compute_max_similarity(emb_rep, emb_kuwana)

    # 結果をDataFrameに追加（表示用は元の質問文を使う）
    soumu_q_texts = df_soumu["質問"].tolist()
    kuwana_q_texts = df_kuwana["質問"].tolist()

    df_rep["総務省_類似度"] = soumu_sim
    df_rep["総務省_最類似質問"] = [soumu_q_texts[i] for i in soumu_idx]
    df_rep["総務省_最類似カテゴリ"] = [df_soumu.iloc[i]["カテゴリ"] for i in soumu_idx]
    df_rep["桑名市_類似度"] = kuwana_sim
    df_rep["桑名市_最類似質問"] = [kuwana_q_texts[i] for i in kuwana_idx]
    df_rep["桑名市_最類似カテゴリ"] = [df_kuwana.iloc[i]["カテゴリ"] for i in kuwana_idx]

    # フィルタリング判定
    df_rep["判定段階"] = ""
    df_rep["最終判定"] = ""
    df_rep["判定根拠"] = ""

    soumu_gray_list = []
    kuwana_gray_list = []

    stats = {
        "total_representative": len(df_rep),
        "soumu_auto_exclude": 0,
        "soumu_llm_exclude": 0,
        "kuwana_auto_exclude": 0,
        "kuwana_llm_exclude": 0,
        "final_shortage": 0,
    }

    for idx, row in df_rep.iterrows():
        s_sim = row["総務省_類似度"]
        k_sim = row["桑名市_類似度"]

        # --- 総務省FAQチェック ---
        if s_sim >= args.soumu_high:
            df_rep.at[idx, "判定段階"] = "総務省FAQ自動除外"
            df_rep.at[idx, "最終判定"] = "除外（総務省カバー）"
            df_rep.at[idx, "判定根拠"] = f"総務省類似度{s_sim:.3f}≧{args.soumu_high}"
            stats["soumu_auto_exclude"] += 1
            continue

        # --- 桑名市FAQチェック ---
        if k_sim >= args.kuwana_high:
            df_rep.at[idx, "判定段階"] = "桑名市FAQ自動除外"
            df_rep.at[idx, "最終判定"] = "除外（桑名市カバー済）"
            df_rep.at[idx, "判定根拠"] = f"桑名市類似度{k_sim:.3f}≧{args.kuwana_high}"
            stats["kuwana_auto_exclude"] += 1
            continue

        # --- グレーゾーン判定 ---
        if args.soumu_low <= s_sim < args.soumu_high:
            soumu_gray_list.append({
                "df_idx": idx,
                "question": row["質問"],
                "zone": "soumu_gray",
                "similar_question": row["総務省_最類似質問"],
                "similarity": s_sim,
                "category": row.get("統一カテゴリ", ""),
            })

        if args.kuwana_low <= k_sim < args.kuwana_high:
            kuwana_gray_list.append({
                "df_idx": idx,
                "question": row["質問"],
                "zone": "kuwana_gray",
                "similar_question": row["桑名市_最類似質問"],
                "similarity": k_sim,
                "category": row.get("統一カテゴリ", ""),
            })

    print(f"  総務省FAQ自動除外: {stats['soumu_auto_exclude']}件")
    print(f"  桑名市FAQ自動除外: {stats['kuwana_auto_exclude']}件")
    print(f"  総務省グレーゾーン: {len(soumu_gray_list)}件")
    print(f"  桑名市グレーゾーン: {len(kuwana_gray_list)}件")

    # ④ LLM判定（グレーゾーン）
    gray_total = len(soumu_gray_list) + len(kuwana_gray_list)
    if gray_total > 0:
        print(f"\n【STEP 4】LLM精密判定（{gray_total}件）")

        if soumu_gray_list:
            print(f"  --- 総務省FAQグレーゾーン: {len(soumu_gray_list)}件 ---")
            soumu_results = llm_judge_batch(soumu_gray_list, api_key)
            for item in soumu_results:
                idx = item["df_idx"]
                if item["llm_verdict"] == "除外":
                    df_rep.at[idx, "判定段階"] = "総務省グレーゾーン→LLM除外"
                    df_rep.at[idx, "最終判定"] = "除外（総務省カバー）"
                    df_rep.at[idx, "判定根拠"] = f"LLM: {item['llm_reason']}"
                    stats["soumu_llm_exclude"] += 1
                else:
                    # 総務省ではカバーされていない → 桑名市チェックへ進む
                    if df_rep.at[idx, "判定段階"] == "":
                        df_rep.at[idx, "判定段階"] = "総務省グレーゾーン→LLM通過"
                        df_rep.at[idx, "判定根拠"] = f"LLM: {item['llm_reason']}"

        if kuwana_gray_list:
            print(f"  --- 桑名市FAQグレーゾーン: {len(kuwana_gray_list)}件 ---")
            kuwana_results = llm_judge_batch(kuwana_gray_list, api_key)
            for item in kuwana_results:
                idx = item["df_idx"]
                # 既に除外済みならスキップ
                if "除外" in str(df_rep.at[idx, "最終判定"]):
                    continue
                if item["llm_verdict"] == "除外":
                    df_rep.at[idx, "判定段階"] = "桑名市グレーゾーン→LLM除外"
                    df_rep.at[idx, "最終判定"] = "除外（桑名市カバー済）"
                    df_rep.at[idx, "判定根拠"] = f"LLM: {item['llm_reason']}"
                    stats["kuwana_llm_exclude"] += 1
                else:
                    if df_rep.at[idx, "判定段階"] == "" or "通過" in str(df_rep.at[idx, "判定段階"]):
                        df_rep.at[idx, "判定段階"] = "桑名市グレーゾーン→LLM通過"
                        if df_rep.at[idx, "判定根拠"] == "":
                            df_rep.at[idx, "判定根拠"] = f"LLM: {item['llm_reason']}"

    else:
        print("\n【STEP 4】グレーゾーンなし → LLM判定スキップ")

    # 未判定のものは「不足」
    for idx, row in df_rep.iterrows():
        if row["最終判定"] == "":
            # 判定段階に応じたラベル付け
            if row["判定段階"] == "":
                s_sim = row["総務省_類似度"]
                k_sim = row["桑名市_類似度"]
                if s_sim < args.soumu_low and k_sim < args.kuwana_low:
                    df_rep.at[idx, "判定段階"] = "自動判定（両方低類似度）"
                elif s_sim < args.soumu_low:
                    df_rep.at[idx, "判定段階"] = "自動判定（総務省低類似度）"
                else:
                    df_rep.at[idx, "判定段階"] = "自動判定"
                df_rep.at[idx, "判定根拠"] = (
                    f"総務省{s_sim:.3f}<{args.soumu_low}, "
                    f"桑名市{k_sim:.3f}<{args.kuwana_low}"
                )
            df_rep.at[idx, "最終判定"] = "不足"

    stats["final_shortage"] = len(df_rep[df_rep["最終判定"] == "不足"])

    print(f"\n【結果サマリー】")
    print(f"  代表質問（入力）:           {stats['total_representative']}件")
    print(f"  総務省FAQ自動除外:          {stats['soumu_auto_exclude']}件")
    print(f"  総務省FAQグレー→LLM除外:    {stats['soumu_llm_exclude']}件")
    print(f"  桑名市FAQ自動除外:          {stats['kuwana_auto_exclude']}件")
    print(f"  桑名市FAQグレー→LLM除外:    {stats['kuwana_llm_exclude']}件")
    print(f"  ──────────────────────────────")
    print(f"  桑名市で不足している質問:    {stats['final_shortage']}件")

    # ⑤ 出力
    print("\n【STEP 5】Excel出力")
    save_results(df_rep, stats, args.output)

    print("\n" + "=" * 60)
    print("処理完了！")
    print(f"  不足質問: {stats['final_shortage']}件")
    print(f"  ファイル: {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()
