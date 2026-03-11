"""
自治体FAQデータ統合 → MMRによる代表質問選出スクリプト（Gemini Embedding API版）

処理フロー:
  ① 指定フォルダ内の全Excelから質問・回答をカテゴリ別に抽出・統合
  ② 質問テキストの正規化・自治体固有質問の除外
  ③ Gemini Embedding APIで全質問をベクトル化（非同期並列処理）
  ④ カテゴリごとにMMRで代表質問を選出
  → 出力: カテゴリ別シートのExcelファイル

使い方:
  export GEMINI_API_KEY="your-api-key"
  python faq_dedup_mmr.py --input ./faq_data --output ./代表質問一覧.xlsx

  # パラメータ調整例
  python faq_dedup_mmr.py -i ./faq_data -o ./output.xlsx --lambda 0.3 --max-ratio 0.5

必要パッケージ:
  pip install pandas openpyxl numpy scikit-learn aiohttp
"""

from __future__ import annotations

from dotenv import load_dotenv
load_dotenv()

import os
import re
import sys
import glob
import asyncio
import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.metrics.pairwise import cosine_similarity

# ============================================================
# 設定（デフォルト値）
# ============================================================
GEMINI_MODEL = "gemini-embedding-001"

# MMR パラメータ
DEFAULT_MMR_LAMBDA = 0.3
DEFAULT_MMR_SCORE_THRESHOLD = -0.5
DEFAULT_MMR_MAX_RATIO = 0.3

# Gemini API 並列処理パラメータ
DEFAULT_BATCH_SIZE = 50
DEFAULT_MAX_CONCURRENT = 5
DEFAULT_RATE_LIMIT_DELAY = 0.5


# ============================================================
# 自治体固有質問フィルタ
# ============================================================
# 自治体名リスト（ファイル名から自動的に追加される + 手動追加分）
EXTRA_CITY_NAMES = {
    # --- 大都市 ---
    "札幌市", "仙台市", "さいたま市", "千葉市", "横浜市",
    "川崎市", "新潟市", "静岡市", "浜松市", "名古屋市",
    "京都市", "大阪市", "神戸市", "広島市", "岡山市",
    "北九州市", "福岡市",
    # --- 中都市 ---
    "旭川市", "盛岡市", "秋田市", "郡山市", "宇都宮市",
    "川越市", "柏市", "八王子市", "金沢市", "長野市",
    "豊橋市", "奈良市", "明石市", "倉敷市", "松山市",
    "高松市", "大分市",
    # --- 小都市 ---
    "函館市", "弘前市", "鶴岡市", "日立市", "鎌倉市",
    "小田原市", "上田市", "津市", "敦賀市", "生駒市", "米子市",
    "呉市", "四万十市", "佐世保市", "延岡市", "那覇市",
    # --- その他 ---
    "堺市", "熊本市", "相模原市",
}


def is_city_specific(question: str, city_name: str) -> bool:
    """質問が特定の自治体固有の内容かどうかを判定"""
    q = question.strip()
    all_city_names = EXTRA_CITY_NAMES | {city_name}

    # 自治体名（○○市）が含まれていれば除外
    for cn in all_city_names:
        if cn in q:
            return True

    # 「市」なしの駅名パターン（「札幌駅」「仙台駅」等）
    for cn in all_city_names:
        base = re.sub(r'[市町村区]$', '', cn)
        if base and len(base) >= 2 and base + "駅" in q:
            return True

    # 広報紙名（「広報かまくら」「広報さっぽろ」等）
    if re.search(r'広報[ぁ-んァ-ン一-鿿]{2,}', q):
        return True
    # 「市政だより」「市民の友」等の固有広報紙名
    if re.search(r'(市政だより|市民の友|掲示板)', q):
        return True

    # 時限的イベント・事業（年号＋固有名、閉館・終了等）
    if re.search(r'(20\d{2}|令和\d)[年度]*.*(?:大会|フェス|祭|イベント|博覧会|万博)', q):
        return True
    if re.search(r'(?:閉館|閉鎖|閉園|廃止|終了).*(?:について|教えて)', q):
        return True

    return False


def is_invalid_question(question: str) -> bool:
    """質問として成立していない無効なテキストを判定する。

    短すぎる文、外国語、単語のみ、見出し等を除外する。
    """
    q = question.strip()

    # 空・極短（8文字未満）
    if len(q) < 8:
        return True

    # 外国語のみ（英語、中国語等）
    if re.match(r'^[a-zA-Z\s\-\.,:;!?\'\"]+$', q):
        return True
    if re.match(r'^[简繁体中文]+$', q):
        return True
    # 「English」「简体中文」「한국어」等の言語名
    if re.match(r'^(English|简体中文|繁體中文|한국어|Tiếng Việt|Português)$', q, re.IGNORECASE):
        return True

    # 「よくある質問」「FAQ」等のページタイトル
    if re.match(r'^(よくある質問|FAQ|ＦＡＱ|Q&A|お問い合わせ|ホーム)$', q):
        return True

    # 疑問形でもなく、短い名詞句のみ（質問になっていない）
    # 例: 「高度地区」「さんかく岡山」「家庭系ごみの有料化」
    if len(q) < 15 and not re.search(r'[？?]|か[。.]?$|たい|ほしい|ですが|のですが|けど|して|どう', q):
        return True

    return False


def strip_city_names(text: str, city_name: str = "") -> str:
    """質問文から自治体名を除去してEmbedding精度を向上させる。

    例:
      "広報さいたま市が届かない" → "広報が届かない"
      "さいたま市のがん検診の費用" → "がん検診の費用"
    """
    all_names = EXTRA_CITY_NAMES | ({city_name} if city_name else set())
    for cn in sorted(all_names, key=len, reverse=True):
        if cn and cn in text:
            text = text.replace(cn, "")
    # 広報紙の固有名も汎用化（「広報かまくら」→「広報」）
    text = re.sub(r'(広報)[ぁ-んァ-ン]{2,}', r'\1', text)
    text = re.sub(r'(広報紙)[「」ぁ-んァ-ン]+[「」]?', r'\1', text)
    text = re.sub(r'^[のはがをにへでと、]+', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# ============================================================
# 統一カテゴリ定義（LLM分類の基準となる20カテゴリ + その他）
# ============================================================
UNIFIED_CATEGORIES = [
    "住民票・戸籍・届出", "健康・医療", "子育て・教育", "税金",
    "環境・ごみ", "福祉・介護", "防災・消防", "保険・年金",
    "住まい・生活", "産業・労働", "市政・行政", "選挙", "相談",
    "施設", "参画・交流", "入札・契約", "観光・文化", "交通",
    "人権", "マイナンバー", "その他"
]

# LLM分類結果のキャッシュ（同じシート名の再分類を防止してAPI節約）
_sheet_category_cache: dict[str, str] = {}

# 質問単位の分類キャッシュ
_question_category_cache: dict[str, str] = {}


def classify_sheet_with_gemini(sheet_name: str, sample_questions: list[str],
                                api_key: str) -> str:
    """Gemini APIを使ってシート名＋質問サンプルから統一カテゴリを自動判定する。

    Args:
        sheet_name: Excelのシート名
        sample_questions: そのシートに含まれる質問のサンプル（3〜5件）
        api_key: Gemini API キー

    Returns:
        UNIFIED_CATEGORIES のいずれか1つ
    """
    # キャッシュヒット → API呼び出し不要
    if sheet_name in _sheet_category_cache:
        cached = _sheet_category_cache[sheet_name]
        print(f"  [CACHE] '{sheet_name}' → '{cached}'")
        return cached

    # カテゴリ一覧文字列
    categories_str = "\n".join(f"- {c}" for c in UNIFIED_CATEGORIES)

    # 質問サンプル文字列（最大5件）
    samples_str = ""
    if sample_questions:
        samples = sample_questions[:5]
        samples_str = "\n".join(f"  - {q}" for q in samples)
        samples_str = f"\n\n含まれる質問の例:\n{samples_str}"

    prompt = f"""あなたは自治体FAQのカテゴリ分類の専門家です。
以下のシート名（と質問サンプル）を見て、最も適切なカテゴリを1つだけ出力してください。

シート名: 「{sheet_name}」{samples_str}

カテゴリ一覧:
{categories_str}

【ルール】
- 上記カテゴリ一覧の中から1つだけ選んでください
- カテゴリ名のみを出力してください（説明や記号は不要）
- どれにも当てはまらない場合は「その他」と出力してください"""

    try:
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model="gemini-3.1-flash-lite-preview",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=50,
            ),
        )

        result = response.text.strip().strip("「」『』\"' ")

        # LLMの応答が定義済みカテゴリに含まれるか検証
        if result not in UNIFIED_CATEGORIES:
            # 部分一致でフォールバック
            for cat in UNIFIED_CATEGORIES:
                if cat in result or result in cat:
                    result = cat
                    break
            else:
                print(f"  [WARN] LLM応答 '{result}' は未知カテゴリ → 'その他'")
                result = "その他"

        _sheet_category_cache[sheet_name] = result
        print(f"  [LLM] '{sheet_name}' → '{result}'")
        return result

    except Exception as e:
        print(f"  [ERROR] LLM分類失敗 '{sheet_name}': {e} → 'その他'")
        _sheet_category_cache[sheet_name] = "その他"
        return "その他"


def classify_questions_batch(questions: list[str], api_key: str,
                              batch_size: int = 30) -> list[str]:
    """複数の質問を一括でカテゴリ分類する（多ジャンル混在シート用）。

    LLM 1回の呼び出しでbatch_size件をまとめて分類し、API呼び出し回数を抑える。
    """
    categories_str = "\n".join(f"- {c}" for c in UNIFIED_CATEGORIES)
    results = ["その他"] * len(questions)

    for start in range(0, len(questions), batch_size):
        batch = questions[start:start + batch_size]
        questions_text = ""
        for k, q in enumerate(batch):
            questions_text += f"{k+1}. {q}\n"

        prompt = f"""あなたは自治体FAQのカテゴリ分類の専門家です。
以下の質問それぞれについて、最も適切なカテゴリを1つ判定してください。

{questions_text}
カテゴリ一覧:
{categories_str}

【出力形式】
番号とカテゴリ名のみを出力してください。
1: 住民票・戸籍・届出
2: 税金
3: 環境・ごみ"""

        try:
            from google import genai
            from google.genai import types

            client = genai.Client(api_key=api_key)
            response = client.models.generate_content(
                model="gemini-3.1-flash-lite-preview",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=1500,
                ),
            )

            for line in response.text.strip().split("\n"):
                line = line.strip()
                if not line:
                    continue
                m = re.match(r'(\d+)\s*[:.：]\s*(.+)', line)
                if m:
                    idx = int(m.group(1)) - 1
                    cat = m.group(2).strip().strip("「」『』\"' ")
                    if cat not in UNIFIED_CATEGORIES:
                        for uc in UNIFIED_CATEGORIES:
                            if uc in cat or cat in uc:
                                cat = uc
                                break
                        else:
                            cat = "その他"
                    if 0 <= start + idx < len(results):
                        results[start + idx] = cat

        except Exception as e:
            print(f"    [ERROR] バッチ分類失敗: {e}")

    return results


def is_mixed_genre_sheet(questions: list[str], sheet_category: str,
                          sample_size: int = 20) -> bool:
    """シート内の質問が多ジャンルに分散しているか簡易判定する。

    キーワードベースで質問のジャンルを推定し、シート分類と異なるジャンルの
    質問が一定割合以上あれば多ジャンル混在と判定する。
    """
    # カテゴリごとのキーワード（簡易判定用）
    CATEGORY_KEYWORDS = {
        "環境・ごみ": ["ごみ", "リサイクル", "分別", "粗大", "カラス", "ハチ"],
        "税金": ["固定資産税", "市民税", "住民税", "法人市民税", "納税", "税証明", "確定申告", "課税"],
        "保険・年金": ["国民健康保険", "国保", "年金", "後期高齢者", "介護保険"],
        "健康・医療": ["予防接種", "検診", "がん検診", "健康診査", "乳幼児健診"],
        "福祉・介護": ["介護", "障害", "障がい", "認知症", "高齢者福祉", "生活保護"],
        "住民票・戸籍・届出": ["住民票", "戸籍", "婚姻届", "離婚届", "出生届", "印鑑登録", "マイナンバーカード"],
        "防災・消防": ["消防", "救急", "避難", "火災", "AED", "防災"],
        "住まい・生活": ["水道", "下水道", "建築", "住宅", "道路", "公園"],
        "子育て・教育": ["保育", "児童", "学童", "幼稚園", "小学校", "就学"],
        "交通": ["バス", "駐車", "駐輪", "交通安全"],
        "マイナンバー": ["マイナンバー", "個人番号", "電子証明書"],
    }

    samples = questions[:min(sample_size, len(questions))]
    if len(samples) < 5:
        return False

    matched_categories = []
    for q in samples:
        for cat, keywords in CATEGORY_KEYWORDS.items():
            if any(kw in q for kw in keywords):
                matched_categories.append(cat)
                break

    if not matched_categories:
        return False

    # ジャンルの種類数が多い場合は混在と判定
    from collections import Counter
    cat_counts = Counter(matched_categories)
    n_genres = len(cat_counts)
    top_genre_ratio = cat_counts.most_common(1)[0][1] / len(matched_categories)

    is_mixed = n_genres >= 4 or (n_genres >= 3 and top_genre_ratio < 0.5)
    return is_mixed


# ============================================================
# ① データ読み込み・統合
# ============================================================
# 読み込みから除外するファイルパターン（代表質問一覧=出力ファイル）
EXCLUDE_FILE_PATTERNS = [
    "代表質問一覧",  # 前回の出力ファイル
]


def load_all_faq(input_dir: str, api_key: str) -> pd.DataFrame:
    all_xlsx = sorted(set(glob.glob(os.path.join(input_dir, "*.xlsx"))))
    # 除外パターンに該当するファイルを除外
    files = []
    excluded = []
    for f in all_xlsx:
        fname = Path(f).name
        if any(pat in fname for pat in EXCLUDE_FILE_PATTERNS):
            excluded.append(fname)
        else:
            files.append(f)

    if not files:
        raise FileNotFoundError(f"Excelファイルが見つかりません: {input_dir}")

    print(f"検出ファイル: {len(files)}件")
    for f in files:
        print(f"  - {Path(f).name}")
    if excluded:
        print(f"除外ファイル: {len(excluded)}件")
        for e in excluded:
            print(f"  ✕ {e}")

    all_rows = []
    skipped_no_question = 0
    skipped_city_specific = 0
    skipped_invalid = 0
    reclassified_count = 0

    for fpath in files:
        city_name = re.sub(r'FAQ.*$', '', Path(fpath).stem).strip() or Path(fpath).stem
        xl = pd.ExcelFile(fpath)

        for sheet_name in xl.sheet_names:
            df = pd.read_excel(fpath, sheet_name=sheet_name)
            if "質問" not in df.columns or "回答" not in df.columns:
                print(f"  [SKIP] {city_name} / {sheet_name}: '質問'または'回答'列なし")
                continue

            # 質問サンプルを取得してLLMに渡す（最大5件）
            sample_qs = []
            if "質問" in df.columns:
                sample_qs = df["質問"].dropna().head(5).tolist()

            unified_category = classify_sheet_with_gemini(sheet_name, sample_qs, api_key)

            # --- 問題1対策: 多ジャンル混在シートの検出 ---
            all_questions_in_sheet = df["質問"].dropna().tolist()
            use_per_question = False
            per_question_categories = {}

            if len(all_questions_in_sheet) >= 10:
                if is_mixed_genre_sheet(all_questions_in_sheet, unified_category):
                    print(f"  [MIXED] {city_name} / {sheet_name}: "
                          f"多ジャンル混在を検出 → 質問単位で再分類 ({len(all_questions_in_sheet)}件)")
                    cats = classify_questions_batch(all_questions_in_sheet, api_key)
                    per_question_categories = {q: c for q, c in zip(all_questions_in_sheet, cats)}
                    use_per_question = True
                    reclassified_count += len(all_questions_in_sheet)

            for _, row in df.iterrows():
                q = str(row.get("質問", "")).strip()
                a = str(row.get("回答", "")).strip()
                sub = str(row.get("サブカテゴリ", "")).strip()

                if not q or q == "nan" or q == "キーワードで検索":
                    skipped_no_question += 1
                    continue

                if is_invalid_question(q):
                    skipped_invalid += 1
                    continue

                if is_city_specific(q, city_name):
                    skipped_city_specific += 1
                    continue

                # 質問単位分類が有効ならそちらを使用
                cat = per_question_categories.get(q, unified_category) if use_per_question else unified_category

                all_rows.append({
                    "自治体": city_name,
                    "元カテゴリ": sheet_name,
                    "統一カテゴリ": cat,
                    "サブカテゴリ": sub if sub != "nan" else "",
                    "質問": q,
                    "回答": a,
                })

    result = pd.DataFrame(all_rows)
    print(f"\n読み込み完了: {len(files)}ファイル, {len(result)}件")
    if skipped_no_question:
        print(f"  無効な質問でスキップ: {skipped_no_question}件")
    if skipped_city_specific:
        print(f"  自治体固有質問で除外: {skipped_city_specific}件")
    if skipped_invalid:
        print(f"  無効な質問文で除外: {skipped_invalid}件")
    if reclassified_count:
        print(f"  多ジャンル混在シートで質問単位再分類: {reclassified_count}件")

    # --- 問題2対策: テンプレートFAQの検出・削減 ---
    result = _reduce_template_faqs(result)

    print("\nカテゴリ別件数:")
    for cat, cnt in result["統一カテゴリ"].value_counts().items():
        print(f"  {cat}: {cnt}件")

    unmapped = result[result["統一カテゴリ"] == "その他"]["元カテゴリ"].unique()
    if len(unmapped) > 0:
        print(f"\n⚠ 'その他'に分類されたシート名:")
        for s in unmapped:
            print(f"    - '{s}'")

    return result


def _reduce_template_faqs(df: pd.DataFrame,
                           max_per_group: int = 3,
                           min_group_size: int = 8) -> pd.DataFrame:
    """同一自治体・同一サブカテゴリから大量に出ているテンプレートFAQを削減する。

    例: 鎌倉市「こどもの家」36件 → 代表3件に削減
        日立市「交流センター」4件 → そのまま（8件未満なので対象外）

    Args:
        df: 全質問DataFrame
        max_per_group: テンプレートグループから残す最大件数
        min_group_size: この件数以上のグループをテンプレートとみなす
    """
    before = len(df)
    groups = df.groupby(["自治体", "サブカテゴリ"])

    rows_to_drop = []
    template_log = []

    for (city, sub), group_df in groups:
        if not sub or sub == "nan" or sub == "":
            continue
        if len(group_df) < min_group_size:
            continue

        # テンプレートFAQ候補: 同一自治体・同一サブカテゴリにmin_group_size件以上
        # → 先頭max_per_group件だけ残し、残りを除去
        excess = group_df.index.tolist()[max_per_group:]
        rows_to_drop.extend(excess)
        template_log.append((city, sub, len(group_df), len(excess)))

    if rows_to_drop:
        df = df.drop(index=rows_to_drop).reset_index(drop=True)
        print(f"\n  テンプレートFAQ削減: {before}件 → {len(df)}件 ({len(rows_to_drop)}件除去)")
        for city, sub, total, dropped in template_log:
            print(f"    {city}/{sub}: {total}件 → {total - dropped}件")

    return df


# ============================================================
# ② テキスト正規化
# ============================================================
def clean_question(text: str) -> str:
    text = re.sub(r'[　\s]+', ' ', text).strip()
    text = re.sub(r'^Q[\.:：\s]*', '', text)
    text = re.sub(r'[？?]+$', '', text).strip()
    return text


def strip_city_names(text: str, city_name: str = "") -> str:
    """質問文から自治体名を除去してEmbedding精度を向上させる。

    例:
      "広報さいたま市が届かない" → "広報が届かない"
      "さいたま市のがん検診の費用と受け方" → "がん検診の費用と受け方"
    """
    all_names = EXTRA_CITY_NAMES | ({city_name} if city_name else set())
    for cn in sorted(all_names, key=len, reverse=True):
        if cn and cn in text:
            text = text.replace(cn, "")
    # 広報紙の固有名も汎用化（「広報かまくら」→「広報」、「広報津」→「広報」）
    text = re.sub(r'(広報)[ぁ-んァ-ン一-鿿]{1,}', r'\1', text)
    # カギ括弧付きの広報紙名を汎用化（「広報つるおか」→ 広報）
    text = re.sub(r'[「『]広報[^」』]*[」』]', '広報', text)
    text = re.sub(r'[「『][^」』]*市[民政][^」』]*[」』]', '広報', text)
    # 広報紙名パターン（「きょうと市民しんぶん」等）
    text = re.sub(r'[「『][ぁ-んァ-ン一-鿿]+[しだ][んよ][ぶり][んー][」』]?', '広報', text)
    text = re.sub(r'^[のはがをにへでと、]+', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# 表記揺れの正規化辞書（Embedding前に適用）
SYNONYM_MAP = [
    (r'国保', '国民健康保険'),
    (r'国民健康保険', '国民健康保険'),  # 統一形を維持
    (r'マイナンバーカード|マイナカード|個人番号カード', 'マイナンバーカード'),
    (r'コンビニエンスストア|コンビニ', 'コンビニ'),
]


def normalize_synonyms(text: str) -> str:
    """表記揺れを統一してEmbedding精度を向上させる"""
    # 「国保」→「国民健康保険」（ただし「国民健康保険」自体はそのまま）
    text = re.sub(r'国保(?!険)', '国民健康保険', text)
    text = re.sub(r'コンビニエンスストア', 'コンビニ', text)
    text = re.sub(r'マイナカード|個人番号カード', 'マイナンバーカード', text)
    return text


def clean_answer(text: str) -> str:
    text = re.sub(r'\$\(function\(\)\s*\{.*', '', text, flags=re.DOTALL)
    text = re.sub(r'<script.*?</script>', '', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'https?://\S+', '', text)
    text = re.sub(r'[　\s]+', ' ', text).strip()
    return text


# ============================================================
# ③ Gemini Embedding API（非同期並列処理）
# ============================================================
def embed_with_gemini(texts: list[str], api_key: str,
                      batch_size: int = DEFAULT_BATCH_SIZE,
                      max_concurrent: int = DEFAULT_MAX_CONCURRENT,
                      rate_limit_delay: float = DEFAULT_RATE_LIMIT_DELAY) -> np.ndarray:
    try:
        import aiohttp
    except ImportError:
        print("ERROR: aiohttp が必要です。  pip install aiohttp")
        sys.exit(1)

    async def _run():
        batches = [texts[i:i+batch_size] for i in range(0, len(texts), batch_size)]
        semaphore = asyncio.Semaphore(max_concurrent)
        total = len(batches)
        print(f"\nGemini Embedding: {len(texts)}件 → {total}バッチ "
              f"(バッチサイズ{batch_size}, 並列数{max_concurrent})")

        async def process_batch(batch_texts, session, batch_id):
            url = (f"https://generativelanguage.googleapis.com/v1beta/"
                   f"models/{GEMINI_MODEL}:batchEmbedContents?key={api_key}")
            body = {"requests": [
                {"model": f"models/{GEMINI_MODEL}",
                 "content": {"parts": [{"text": t}]},
                 "taskType": "CLUSTERING",
                 "outputDimensionality": 1536}
                for t in batch_texts
            ]}

            for attempt in range(3):
                async with semaphore:
                    try:
                        async with session.post(url, json=body) as resp:
                            if resp.status == 429:
                                wait = float(resp.headers.get("Retry-After", 5))
                                print(f"  [RATE LIMIT] バッチ{batch_id}: "
                                      f"{wait}秒待機 (リトライ{attempt+1}/3)")
                                await asyncio.sleep(wait)
                                continue
                            elif resp.status != 200:
                                error = await resp.text()
                                print(f"  [ERROR] バッチ{batch_id}: "
                                      f"status={resp.status}, {error[:200]}")
                                return [None] * len(batch_texts)
                            data = await resp.json()
                            return [e.get("values")
                                    for e in data.get("embeddings", [])]
                    except Exception as e:
                        print(f"  [ERROR] バッチ{batch_id}: {e}")
                        if attempt < 2:
                            await asyncio.sleep(2)
                    finally:
                        await asyncio.sleep(rate_limit_delay)
            return [None] * len(batch_texts)

        async with aiohttp.ClientSession() as session:
            tasks = [process_batch(b, session, i)
                     for i, b in enumerate(batches)]
            results = await asyncio.gather(*tasks)

        all_embs, failed, dim = [], 0, None
        for batch_result in results:
            for emb in batch_result:
                if emb is None:
                    failed += 1
                    all_embs.append(None)
                else:
                    if dim is None:
                        dim = len(emb)
                    all_embs.append(np.array(emb))

        if dim is None:
            raise RuntimeError(
                "全Embeddingに失敗しました。"
                "APIキーとネットワーク接続を確認してください。")

        for i in range(len(all_embs)):
            if all_embs[i] is None:
                all_embs[i] = np.zeros(dim)

        if failed:
            print(f"  ⚠ {failed}件のEmbeddingに失敗（ゼロベクトルで代替）")

        print(f"Embedding完了: {len(all_embs)}件, 次元数: {dim}")
        return np.array(all_embs)

    return asyncio.run(_run())


# ============================================================
# ④ MMR（Maximal Marginal Relevance）
# ============================================================
def mmr_select(embeddings, indices,
               lam=DEFAULT_MMR_LAMBDA,
               score_threshold=DEFAULT_MMR_SCORE_THRESHOLD,
               max_ratio=DEFAULT_MMR_MAX_RATIO,
               dup_sim_threshold=0.92,
               cities=None):
    """MMR選出（改良版: cross-city relevance + 高類似度ペア強制排除）

    Args:
        dup_sim_threshold: この値以上の類似度を持つ質問は強制排除。
        cities: 各インデックスに対応する自治体名のリスト。
                指定時、複数自治体に共通する質問を優先する（cross-city relevance）。
    """
    if len(indices) <= 1:
        return indices

    sub = embeddings[indices]
    n = len(indices)
    max_select = max(1, int(n * max_ratio))

    # 事前に全ペアの類似度行列を計算
    sim_matrix = cosine_similarity(sub)  # (n, n)

    # --- relevance計算 ---
    if cities is not None and len(set(cities)) >= 2:
        # Cross-City Relevance: 他の自治体の質問との平均最大類似度
        # 各質問iについて、自治体ごとに最も類似する質問との類似度を求め、
        # その自治体平均を取ることで「どれだけ多くの自治体で共通する話題か」を測る
        city_arr = np.array(cities)
        unique_cities = list(set(cities))
        relevance = np.zeros(n)

        for i in range(n):
            my_city = city_arr[i]
            other_city_max_sims = []
            for uc in unique_cities:
                if uc == my_city:
                    continue
                mask = city_arr == uc
                if not np.any(mask):
                    continue
                max_sim = sim_matrix[i, mask].max()
                other_city_max_sims.append(max_sim)
            if other_city_max_sims:
                relevance[i] = np.mean(other_city_max_sims)
            else:
                relevance[i] = 0.0
    else:
        # フォールバック: 従来の重心ベース
        centroid = sub.mean(axis=0, keepdims=True)
        relevance = cosine_similarity(sub, centroid).flatten()

    selected = []
    remaining = set(range(n))

    # 最も関連性の高いものを最初に選択
    first = int(np.argmax(relevance))
    selected.append(first)
    remaining.discard(first)

    # 各候補の「選択済みとの最大類似度」を追跡
    max_sim_to_selected = sim_matrix[:, first].copy()

    remaining_arr = np.array(list(remaining))

    while len(remaining) > 0 and len(selected) < max_select:
        r_idx = remaining_arr

        # 強制排除: 選択済みとの最大類似度が閾値以上の候補を除外
        not_dup_mask = max_sim_to_selected[r_idx] < dup_sim_threshold
        if not np.any(not_dup_mask):
            break

        scores = lam * relevance[r_idx] - (1 - lam) * max_sim_to_selected[r_idx]
        scores[~not_dup_mask] = -np.inf

        best_pos = int(np.argmax(scores))
        best_score = scores[best_pos]

        if best_score < score_threshold or best_score == -np.inf:
            break

        best_idx = r_idx[best_pos]
        selected.append(int(best_idx))
        remaining.discard(int(best_idx))

        remaining_arr = np.delete(remaining_arr, best_pos)
        np.maximum(max_sim_to_selected, sim_matrix[:, best_idx], out=max_sim_to_selected)

    return [indices[i] for i in selected]


# ============================================================
# ⑤ 地域固有質問の最終フィルタ（LLMベース）
# ============================================================
def _filter_region_specific(df: pd.DataFrame, api_key: str,
                             batch_size: int = 40) -> tuple[pd.DataFrame, int]:
    """MMR選出後の質問に対し、LLMで地域固有質問を判定して除去する。

    「質問文中に特定の地名・施設固有名・路線固有名が含まれているか」のみを判定。
    質問の内容やテーマで判定するのではなく、固有名詞の有無で判定する。
    """
    questions = df["質問"].tolist()

    is_regional = [False] * len(questions)

    for start in range(0, len(questions), batch_size):
        batch_qs = questions[start:start + batch_size]

        questions_text = ""
        for k, q in enumerate(batch_qs):
            questions_text += f"{k+1}. {q}\n"

        prompt = f"""あなたは日本語テキストから固有名詞を検出する専門家です。
以下の自治体FAQ質問リストの中から、**質問文中にその地域でしか通じない固有名詞**が含まれているものだけを選んでください。

{questions_text}
【「地域固有」と判定する条件 — 以下のいずれかに該当する場合のみ】
① 特定の地名・地区名・通り名が含まれる（例: 「須磨海水浴場」「都祁行政センター」「八条・大安寺周辺地区」）
② 特定の鉄道路線名・有料道路名が含まれる（例: 「西神・山手線」「六甲有料道路」「宇都宮LRT」「ゆいレール」）
③ 特定の施設の固有名が含まれる（例: 「こども本の森 神戸」「シルバーカレッジ」「UNITY」）
④ 特定の地域だけの固有制度名・イベント名が含まれる（例: 「認知症神戸モデル」「灘の酒蔵謎解き探訪」「ワケトンカレンダー」）
⑤ 特定の県名が含まれる（例: 「奈良県自転車条例」「兵庫県収入証紙」）

【「汎用」と判定する（＝除外しない）もの — 非常に重要】
- 質問のテーマが一般的であれば、出典の自治体に関係なく「汎用」です
- 以下はすべて「汎用」です。絶対に地域固有と判定しないでください:
  ・全国共通の制度: 児童手当、国民健康保険、介護保険、生活保護、住民票、戸籍、固定資産税、粗大ごみ、etc.
  ・どの自治体にもある質問: ごみの分別方法、転入転出の手続き、証明書の取得方法、防災訓練、ハザードマップ、etc.
  ・一般的な行政サービス: 図書館、保育所、学童保育、市営住宅、消防署見学、パブリックコメント、etc.
  ・一般的な社会問題: DV相談、野良猫対策、空き家対策、騒音問題、不審者対応、etc.
- 「どこで」「いつ」「いくら」等を聞いていても、質問自体はどの自治体でも成り立つなら「汎用」です

【出力形式】
地域固有と判定した質問の番号のみをカンマ区切りで出力してください。
1件もなければ「なし」と出力してください。"""

        try:
            from google import genai
            from google.genai import types

            client = genai.Client(api_key=api_key)
            response = client.models.generate_content(
                model="gemini-3.1-flash-lite-preview",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=500,
                ),
            )

            text = response.text.strip()
            if text != "なし" and text != "該当なし":
                for num_str in re.findall(r'\d+', text):
                    idx = int(num_str) - 1
                    if 0 <= idx < len(batch_qs):
                        is_regional[start + idx] = True

        except Exception as e:
            print(f"    [ERROR] バッチ判定失敗: {e}")

        import time
        time.sleep(1)

    # 地域固有と判定された質問を除去
    removed_indices = [i for i, flag in enumerate(is_regional) if flag]
    if removed_indices:
        removed_qs = df.iloc[removed_indices]
        print(f"  地域固有と判定: {len(removed_indices)}件")
        for _, row in removed_qs.iterrows():
            print(f"    除去: [{row['統一カテゴリ']}] {row['質問'][:60]} ({row['自治体']})")
        df = df.drop(index=removed_indices).reset_index(drop=True)

    return df, len(removed_indices)


# ============================================================
# Excel出力
# ============================================================
def save_to_excel(df, category_stats, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side("thin"), right=Side("thin"),
                top=Side("thin"), bottom=Side("thin"))

    # --- サマリーシート ---
    ws = wb.create_sheet("サマリー", 0)
    for ci, h in enumerate(["カテゴリ", "元の質問数", "代表質問数", "削減率"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for ri, cat in enumerate(sorted(category_stats.keys()), 2):
        s = category_stats[cat]
        red = f"{(1-s['after']/s['before'])*100:.0f}%" if s['before'] > 0 else "-"
        for ci, v in enumerate([cat, s["before"], s["after"], red], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = cf, tb

    tr = len(category_stats) + 2
    tb_before = sum(s["before"] for s in category_stats.values())
    tb_after = sum(s["after"] for s in category_stats.values())
    red_t = f"{(1-tb_after/tb_before)*100:.0f}%" if tb_before > 0 else "-"
    bf = Font(name="Arial", bold=True, size=11)
    for ci, v in enumerate(["合計", tb_before, tb_after, red_t], 1):
        c = ws.cell(row=tr, column=ci, value=v)
        c.font, c.border = bf, tb

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    # --- カテゴリ別シート ---
    headers = ["No.", "統一カテゴリ", "サブカテゴリ",
               "質問", "回答（参考）", "出典自治体"]
    for cat in sorted(df["統一カテゴリ"].unique()):
        cdf = df[df["統一カテゴリ"] == cat].reset_index(drop=True)
        sn = re.sub(r'[\\/*?\[\]:]', '', cat)[:31]
        ws = wb.create_sheet(title=sn)

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

        for ri, (_, row) in enumerate(cdf.iterrows(), 2):
            ans = str(row["回答"])
            if len(ans) > 500:
                ans = ans[:500] + "..."
            vals = [ri-1, row["統一カテゴリ"], row.get("サブカテゴリ", ""),
                    row["質問"], ans, row["自治体"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font, c.alignment, c.border = cf, ca, tb

        for ci, w in enumerate([6, 18, 18, 50, 60, 12], 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
        ws.auto_filter.ref = f"A1:F{len(cdf)+1}"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n出力完了: {output_path}")


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="自治体FAQ 代表質問抽出ツール（Gemini API版）")
    parser.add_argument("--input", "-i", required=True,
                        help="FAQのExcelファイル格納ディレクトリ")
    parser.add_argument("--output", "-o", default="./代表質問一覧.xlsx",
                        help="出力Excelファイルパス")
    parser.add_argument("--api-key",
                        help="Gemini APIキー（環境変数 GEMINI_API_KEY でも可）")
    parser.add_argument("--lambda", dest="mmr_lambda", type=float,
                        default=DEFAULT_MMR_LAMBDA,
                        help=f"MMRのλ値 (default: {DEFAULT_MMR_LAMBDA})")
    parser.add_argument("--max-ratio", type=float, default=DEFAULT_MMR_MAX_RATIO,
                        help=f"カテゴリあたり最大選出比率 (default: {DEFAULT_MMR_MAX_RATIO})")
    parser.add_argument("--threshold", type=float, default=DEFAULT_MMR_SCORE_THRESHOLD,
                        help=f"MMR打ち切り閾値 (default: {DEFAULT_MMR_SCORE_THRESHOLD})")
    parser.add_argument("--dup-threshold", type=float, default=0.92,
                        help="MMR内の重複排除類似度閾値 (default: 0.92)")
    parser.add_argument("--intra-threshold", type=float, default=0.85,
                        help="STEP4-2 同カテゴリ内LLM重複判定の類似度閾値 (default: 0.85)")
    parser.add_argument("--cross-threshold", type=float, default=0.87,
                        help="STEP4-2 カテゴリ間LLM重複判定の類似度閾値 (default: 0.87)")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE,
                        help=f"APIバッチサイズ (default: {DEFAULT_BATCH_SIZE})")
    parser.add_argument("--max-concurrent", type=int, default=DEFAULT_MAX_CONCURRENT,
                        help=f"最大並列数 (default: {DEFAULT_MAX_CONCURRENT})")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: Gemini APIキーが必要です。")
        print("  --api-key YOUR_KEY または $env:GEMINI_API_KEY='YOUR_KEY'")
        sys.exit(1)

    print("=" * 60)
    print("自治体FAQ 代表質問抽出ツール（Gemini API版）")
    print(f"  入力: {args.input}")
    print(f"  出力: {args.output}")
    print(f"  MMR: λ={args.mmr_lambda}, 最大比率={args.max_ratio}, "
          f"閾値={args.threshold}, 重複閾値={args.dup_threshold}")
    print(f"  STEP4-2: 同カテゴリ内≥{args.intra_threshold}, "
          f"カテゴリ間≥{args.cross_threshold}")
    print(f"  API: バッチ={args.batch_size}, 並列={args.max_concurrent}")
    print("=" * 60)

    # ① 読み込み
    print("\n【STEP 1】データ読み込み（自治体固有質問を自動除外）")
    df = load_all_faq(args.input, api_key)
    if len(df) == 0:
        print("有効な質問データがありません。")
        sys.exit(1)

    # ② 正規化
    print("\n【STEP 2】テキスト正規化")
    df["質問_clean"] = df["質問"].apply(clean_question)
    df["回答"] = df["回答"].apply(clean_answer)

    # Embedding用: 自治体名を除去（「広報○○市」→「広報」で統一）
    df["質問_embed"] = df.apply(
        lambda row: normalize_synonyms(
            strip_city_names(row["質問_clean"], row.get("自治体", ""))
        ),
        axis=1,
    )
    n_changed = len(df[df["質問_clean"] != df["質問_embed"]])
    if n_changed > 0:
        print(f"  自治体名除去: {n_changed}件が変更")
        for _, row in df[df["質問_clean"] != df["質問_embed"]].head(3).iterrows():
            print(f"    {row['質問_clean'][:35]} → {row['質問_embed'][:35]}")

    before = len(df)
    df = df.drop_duplicates(subset=["質問_embed"], keep="first").reset_index(drop=True)
    print(f"  完全一致除去: {before} → {len(df)}件（{before-len(df)}件除去）")

    # ③ Embedding
    print("\n【STEP 3】Gemini Embedding（非同期並列処理）")
    embeddings = embed_with_gemini(
        df["質問_embed"].tolist(),
        api_key,
        batch_size=args.batch_size,
        max_concurrent=args.max_concurrent,
    )

    # ④ MMR（Cross-City Relevance: 複数自治体に共通する質問を優先）
    print("\n【STEP 4】MMRによる代表質問選出（cross-city relevance）")
    stats, selected_indices = {}, []
    for cat in sorted(df["統一カテゴリ"].unique()):
        idxs = df[df["統一カテゴリ"] == cat].index.tolist()
        if not idxs:
            continue
        # カテゴリ内の各質問の出典自治体リストを取得
        cat_cities = df.loc[idxs, "自治体"].tolist()
        sel = mmr_select(embeddings, idxs,
                         lam=args.mmr_lambda,
                         score_threshold=args.threshold,
                         max_ratio=args.max_ratio,
                         dup_sim_threshold=args.dup_threshold,
                         cities=cat_cities)
        selected_indices.extend(sel)
        stats[cat] = {"before": len(idxs), "after": len(sel)}
        print(f"  {cat}: {len(idxs)}件 → {len(sel)}件")

    result = df.loc[selected_indices].reset_index(drop=True)
    result_emb = embeddings[selected_indices]
    total_before, total_after = len(df), len(result)
    print(f"\n合計: {total_before}件 → {total_after}件"
          f"（{total_after/total_before*100:.1f}%）")

    # ④-2 全体の重複除去（同カテゴリ内 + カテゴリ間、コサイン類似度 + LLM判定）
    print("\n【STEP 4-2】残存する重複質問を除去（同カテゴリ内 + カテゴリ間、LLMハイブリッド方式）")
    sim = cosine_similarity(result_emb)
    np.fill_diagonal(sim, 0)

    # Phase 1: コサイン類似度が高い候補ペアを抽出（同カテゴリ内も対象に）
    INTRA_THRESHOLD = args.intra_threshold   # 同カテゴリ内: MMRで漏れた重複を捕捉
    CROSS_THRESHOLD = args.cross_threshold   # カテゴリ間
    candidates = []
    for i in range(len(result)):
        for j in range(i + 1, len(result)):
            same_cat = result.iloc[i]["統一カテゴリ"] == result.iloc[j]["統一カテゴリ"]
            threshold = INTRA_THRESHOLD if same_cat else CROSS_THRESHOLD
            if sim[i, j] >= threshold:
                candidates.append((i, j, sim[i, j]))

    print(f"  候補ペア: {len(candidates)}件（同カテゴリ内≥{INTRA_THRESHOLD}, カテゴリ間≥{CROSS_THRESHOLD}）")

    if not candidates:
        print("  カテゴリ間の重複候補なし")
    else:
        # Phase 2: LLMで非同期並列バッチ判定
        import aiohttp

        LLM_BATCH_SIZE = 50
        MAX_CONCURRENT = 5
        LLM_MODEL = "gemini-3.1-flash-lite-preview"

        llm_results = {}  # (i, j) -> True(統合可能) / False(別質問)

        # バッチごとのプロンプトを事前生成
        batches = []
        for batch_start in range(0, len(candidates), LLM_BATCH_SIZE):
            batch = candidates[batch_start:batch_start + LLM_BATCH_SIZE]
            pairs_text = ""
            for k, (i, j, s) in enumerate(batch):
                same_cat_flag = "【同カテゴリ】" if result.iloc[i]['統一カテゴリ'] == result.iloc[j]['統一カテゴリ'] else ""
                pairs_text += (
                    f"ペア{k+1}: {same_cat_flag}\n"
                    f"  A[{result.iloc[i]['統一カテゴリ']}]: {result.iloc[i]['質問']}\n"
                    f"  B[{result.iloc[j]['統一カテゴリ']}]: {result.iloc[j]['質問']}\n\n"
                )
            prompt = f"""あなたは自治体FAQの重複チェックの専門家です。
以下のペアそれぞれについて、2つの質問が「同じ回答で対応できる重複質問」か「別々の回答が必要な別質問」かを判定してください。

{pairs_text}
【判定基準】
- 問い合わせ内容が実質的に同じで、同じ回答で対応できる → 「統合」
- 言い回しが違うだけで聞いていることが同じ → 「統合」
  例: 「職場の保険に入ったので国保を辞めたい」と「就職したので国民健康保険をやめる手続きは」→ 統合
  例: 「粗大ごみの出し方」と「大型ごみの処分方法」→ 統合
  例: 「広報○○が届かない」と「広報△△が届かない」→ 統合（自治体名が違うだけ）
- サブトピックや対象者が異なり、別の回答が必要 → 「別」
  例: 「国民健康保険の加入手続き」と「国民健康保険料の納付方法」→ 別
  例: 「介護認定の申請先」と「介護保険料の計算方法」→ 別

【出力形式】
ペアごとに「番号: 統合 または 別」のみを出力してください。
1: 統合
2: 別
3: 統合"""
            batches.append((batch, prompt))

        print(f"  LLMバッチ数: {len(batches)}（バッチサイズ{LLM_BATCH_SIZE}, 並列{MAX_CONCURRENT}）")

        async def _run_llm_dedup():
            sem = asyncio.Semaphore(MAX_CONCURRENT)
            completed = [0]

            async def process_one(session, batch, prompt, batch_id):
                url = (f"https://generativelanguage.googleapis.com/v1beta/"
                       f"models/{LLM_MODEL}:generateContent?key={api_key}")
                body = {
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {
                        "temperature": 0.0,
                        "maxOutputTokens": 1000,
                    }
                }
                async with sem:
                    for attempt in range(3):
                        try:
                            async with session.post(url, json=body) as resp:
                                if resp.status == 429:
                                    await asyncio.sleep(2 ** attempt)
                                    continue
                                data = await resp.json()
                                text = data["candidates"][0]["content"]["parts"][0]["text"]

                                for line in text.strip().split("\n"):
                                    line = line.strip()
                                    if not line:
                                        continue
                                    m = re.match(r'(\d+)\s*[:.：]\s*(統合|別)', line)
                                    if m:
                                        idx = int(m.group(1)) - 1
                                        if 0 <= idx < len(batch):
                                            i, j, s = batch[idx]
                                            llm_results[(i, j)] = (m.group(2) == "統合")
                                break
                        except Exception as e:
                            if attempt == 2:
                                print(f"  [ERROR] バッチ{batch_id}: {e}")
                                for i, j, s in batch:
                                    llm_results[(i, j)] = (s >= 0.95)
                            else:
                                await asyncio.sleep(1)

                    completed[0] += 1
                    if completed[0] % 5 == 0 or completed[0] == len(batches):
                        print(f"  LLM判定: {completed[0]}/{len(batches)}バッチ完了")

            async with aiohttp.ClientSession() as session:
                tasks = [process_one(session, b, p, idx)
                         for idx, (b, p) in enumerate(batches)]
                await asyncio.gather(*tasks)

        asyncio.run(_run_llm_dedup())

        # Phase 3: LLMが「統合」と判断したペアを除去
        removed = set()
        cross_dup_log = []

        # 類似度が高い順にソートして処理
        for i, j, s in sorted(candidates, key=lambda x: -x[2]):
            if i in removed or j in removed:
                continue
            is_dup = llm_results.get((i, j), False)
            if is_dup:
                removed.add(j)
                cross_dup_log.append((
                    result.iloc[j]["質問"][:50],
                    result.iloc[j]["統一カテゴリ"],
                    result.iloc[i]["質問"][:50],
                    result.iloc[i]["統一カテゴリ"],
                    s,
                ))

        if cross_dup_log:
            intra = sum(1 for _, cat_rm, _, cat_kp, _ in cross_dup_log if cat_rm == cat_kp)
            cross = len(cross_dup_log) - intra
            print(f"\n  LLM判定で{len(cross_dup_log)}件の重複を除去（同カテゴリ内: {intra}件, カテゴリ間: {cross}件）:")
            for q_rm, cat_rm, q_kp, cat_kp, s in cross_dup_log[:15]:
                print(f"    除去: [{cat_rm}] {q_rm}")
                print(f"    残す: [{cat_kp}] {q_kp} (類似度={s:.3f})")
            if len(cross_dup_log) > 15:
                print(f"    ... 他{len(cross_dup_log) - 15}件")

            keep_mask = [i for i in range(len(result)) if i not in removed]
            result = result.iloc[keep_mask].reset_index(drop=True)

            for cat in stats:
                stats[cat]["after"] = len(result[result["統一カテゴリ"] == cat])

            print(f"  {total_after}件 → {len(result)}件")
            total_after = len(result)
        else:
            print("  LLM判定で重複なし")

    # ⑤ 地域固有質問の最終フィルタ（LLMで汎用性を判定）
    print("\n【STEP 5】地域固有質問の除去（LLM汎用性判定）")
    result, n_removed = _filter_region_specific(result, api_key)
    if n_removed > 0:
        for cat in stats:
            stats[cat]["after"] = len(result[result["統一カテゴリ"] == cat])
        total_after = len(result)
        print(f"  → {n_removed}件除去、残り{total_after}件")
    else:
        print("  地域固有質問なし")

    # ⑥ 出力
    print("\n【STEP 6】Excel出力")
    save_to_excel(result, stats, args.output)

    print("\n" + "=" * 60)
    print("処理完了！")
    print(f"  入力: {total_before}件 → 出力: {total_after}件")
    print(f"  ファイル: {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()