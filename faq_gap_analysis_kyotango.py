# -*- coding: utf-8 -*-
"""
FAQ ギャップ分析スクリプト（Gemini Embedding API + LLM判定）
京丹後市バージョン

代表質問から、総務省FAQに載っている質問を除外し、
京丹後市FAQに不足している質問を提案する。

処理フロー:
  ① 3つのExcel（代表質問・総務省FAQ・対象自治体FAQ）を読み込み
  ② Gemini Embedding APIで全質問をベクトル化
  ③ コサイン類似度で候補ペアを抽出
  ④ 微妙なライン（類似度0.75〜0.88）はLLMで最終判定
  ⑤ 結果をExcel出力

使い方:
  # .envファイルにGEMINI_API_KEYを設定（推奨）
  # GEMINI_API_KEY=AIzaSy...
  python faq_gap_analysis_kyotango.py --daihyo 代表質問一覧50件v9.xlsx --soumu 総務省FAQ.xlsx --city 京丹後市FAQ.xlsx --output 京丹後市FAQ提案.xlsx
"""

from __future__ import annotations

from dotenv import load_dotenv
load_dotenv()

import os
import re
import sys
import asyncio
import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.metrics.pairwise import cosine_similarity

# ============================================================
# 設定
# ============================================================
GEMINI_EMBED_MODEL = "gemini-embedding-001"
LLM_MODEL = "gemini-3.1-flash-lite-preview"
EMBED_BATCH_SIZE = 50
EMBED_MAX_CONCURRENT = 5
EMBED_RATE_LIMIT_DELAY = 0.5
EMBED_DIM = 1536

# 類似度閾値
AUTO_MATCH_THRESHOLD = 0.88     # これ以上 → 自動的に「類似」と判定
LLM_JUDGE_THRESHOLD = 0.75     # これ以上かつ AUTO未満 → LLMで判定
# LLM_JUDGE_THRESHOLD 未満 → 自動的に「非類似」と判定

LLM_BATCH_SIZE = 30             # LLM判定のバッチサイズ


# ============================================================
# ① データ読み込み
# ============================================================
def load_questions(path: str, skip_summary: bool = True) -> list[dict]:
    xl = pd.ExcelFile(path)
    rows = []
    for sheet in xl.sheet_names:
        if skip_summary and sheet == "サマリー":
            continue
        df = pd.read_excel(path, sheet_name=sheet)
        q_col = "質問" if "質問" in df.columns else None
        if not q_col:
            continue

        # カテゴリ列の自動検出（京丹後市FAQ対応）
        # 「大カテゴリ」列があればそれを使い、なければシート名をカテゴリとする
        has_main_cat = "大カテゴリ" in df.columns

        for _, row in df.iterrows():
            q = str(row.get(q_col, "")).strip()
            if q and q != "nan":
                # カテゴリの決定
                if has_main_cat:
                    category = str(row.get("大カテゴリ", "")).strip()
                    if not category or category == "nan":
                        category = sheet
                else:
                    category = sheet

                rows.append({
                    "category": category,
                    "subcategory": str(row.get("サブカテゴリ", "")),
                    "question": q,
                    "answer": str(
                        row.get("回答", row.get("回答（参考）", ""))
                    )[:500],
                    "source": str(row.get("出典自治体", "")),
                })
    return rows


# ============================================================
# ② Gemini Embedding API（非同期並列処理）
# ============================================================
def embed_with_gemini(texts: list[str], api_key: str) -> np.ndarray:
    import aiohttp

    async def _run():
        batches = [
            texts[i:i + EMBED_BATCH_SIZE]
            for i in range(0, len(texts), EMBED_BATCH_SIZE)
        ]
        semaphore = asyncio.Semaphore(EMBED_MAX_CONCURRENT)
        total = len(batches)
        print(f"  Embedding: {len(texts)}件 → {total}バッチ")

        async def process_batch(batch_texts, session, batch_id):
            url = (
                f"https://generativelanguage.googleapis.com/v1beta/"
                f"models/{GEMINI_EMBED_MODEL}:batchEmbedContents"
                f"?key={api_key}"
            )
            body = {
                "requests": [
                    {
                        "model": f"models/{GEMINI_EMBED_MODEL}",
                        "content": {"parts": [{"text": t}]},
                        "taskType": "SEMANTIC_SIMILARITY",
                        "outputDimensionality": EMBED_DIM,
                    }
                    for t in batch_texts
                ]
            }
            for attempt in range(3):
                async with semaphore:
                    try:
                        async with session.post(url, json=body) as resp:
                            if resp.status == 429:
                                wait = float(
                                    resp.headers.get("Retry-After", 5)
                                )
                                print(
                                    f"  [RATE LIMIT] バッチ{batch_id}: "
                                    f"{wait}秒待機"
                                )
                                await asyncio.sleep(wait)
                                continue
                            elif resp.status != 200:
                                error = await resp.text()
                                print(
                                    f"  [ERROR] バッチ{batch_id}: "
                                    f"status={resp.status}, {error[:200]}"
                                )
                                return [None] * len(batch_texts)
                            data = await resp.json()
                            return [
                                e.get("values")
                                for e in data.get("embeddings", [])
                            ]
                    except Exception as e:
                        print(f"  [ERROR] バッチ{batch_id}: {e}")
                        if attempt < 2:
                            await asyncio.sleep(2)
                    finally:
                        await asyncio.sleep(EMBED_RATE_LIMIT_DELAY)
            return [None] * len(batch_texts)

        async with aiohttp.ClientSession() as session:
            tasks = [
                process_batch(b, session, i)
                for i, b in enumerate(batches)
            ]
            results = await asyncio.gather(*tasks)

        all_embs, failed, dim = [], 0, None
        for batch_result in results:
            for emb in batch_result:
                if emb is None:
                    failed += 1
                    all_embs.append(None)
                else:
                    if dim is None:
                        dim = len(emb)
                    all_embs.append(np.array(emb))

        if dim is None:
            raise RuntimeError("全Embeddingに失敗しました。APIキーを確認してください。")

        for i in range(len(all_embs)):
            if all_embs[i] is None:
                all_embs[i] = np.zeros(dim)

        if failed:
            print(f"  ⚠ {failed}件のEmbeddingに失敗（ゼロベクトルで代替）")

        print(f"  完了: {len(all_embs)}件, 次元数: {dim}")
        return np.array(all_embs)

    return asyncio.run(_run())


# ============================================================
# ④ LLM判定（微妙な類似度ペアを一括判定）
# ============================================================
def llm_judge_pairs(
    pairs: list[tuple[str, str, float]],
    api_key: str,
    context: str = "総務省FAQ",
) -> dict[int, bool]:
    """
    質問ペアのリストをLLMに渡し、「同じ内容か」を判定。

    Args:
        pairs: [(query_question, target_question, similarity), ...]
        api_key: Gemini API key
        context: 判定対象の説明（ログ用）

    Returns:
        {pair_index: True(同内容) or False(別内容)}
    """
    import aiohttp

    results: dict[int, bool] = {}

    batches = []
    for start in range(0, len(pairs), LLM_BATCH_SIZE):
        batch = pairs[start:start + LLM_BATCH_SIZE]
        batch_indices = list(range(start, start + len(batch)))

        pairs_text = ""
        for k, (q1, q2, sim) in enumerate(batch):
            pairs_text += (
                f"ペア{k+1}:\n"
                f"  A: {q1}\n"
                f"  B: {q2}\n\n"
            )

        prompt = f"""あなたは自治体FAQの類似判定の専門家です。
以下のペアそれぞれについて、AとBが「同じ質問内容（同じ回答で対応できる）」か「別の質問内容」かを判定してください。

{pairs_text}
【判定基準】
- 聞いていることの本質が同じで、同じ回答で対応できる → 「同」
- 表現は似ているが、聞いていることが違う → 「別」
  例: 「固定資産税の評価額について」と「固定資産税の評価額とは何ですか」→ 同
  例: 「国民年金の免除申請」と「国民年金の加入手続き」→ 別
  例: 「粗大ごみの出し方」と「粗大ごみの申し込み方法」→ 同

【出力形式】
番号と判定のみを出力してください。
1: 同
2: 別
3: 同"""

        batches.append((batch_indices, batch, prompt))

    if not batches:
        return results

    print(f"  LLM判定: {len(pairs)}ペア → {len(batches)}バッチ（{context}）")

    async def _run():
        sem = asyncio.Semaphore(5)
        completed = [0]

        async def process_one(session, batch_indices, batch, prompt):
            url = (
                f"https://generativelanguage.googleapis.com/v1beta/"
                f"models/{LLM_MODEL}:generateContent?key={api_key}"
            )
            body = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.0,
                    "maxOutputTokens": 1000,
                },
            }
            async with sem:
                for attempt in range(3):
                    try:
                        async with session.post(url, json=body) as resp:
                            if resp.status == 429:
                                await asyncio.sleep(3 * (attempt + 1))
                                continue
                            data = await resp.json()
                            text = (
                                data["candidates"][0]["content"]
                                ["parts"][0]["text"]
                            )
                            for line in text.strip().split("\n"):
                                line = line.strip()
                                if not line:
                                    continue
                                m = re.match(
                                    r'(\d+)\s*[:.：]\s*(同|別)', line
                                )
                                if m:
                                    idx = int(m.group(1)) - 1
                                    if 0 <= idx < len(batch):
                                        global_idx = batch_indices[idx]
                                        results[global_idx] = (
                                            m.group(2) == "同"
                                        )
                            break
                    except Exception as e:
                        if attempt == 2:
                            print(f"    [ERROR] LLMバッチ: {e}")
                        else:
                            await asyncio.sleep(2)

                completed[0] += 1
                if completed[0] % 5 == 0 or completed[0] == len(batches):
                    print(
                        f"    LLM: {completed[0]}/{len(batches)}バッチ完了"
                    )

        async with aiohttp.ClientSession() as session:
            tasks = [
                process_one(session, bi, b, p)
                for bi, b, p in batches
            ]
            await asyncio.gather(*tasks)

    asyncio.run(_run())
    return results


# ============================================================
# ③⑤ メインマッチングロジック
# ============================================================
def match_questions(
    query_embs: np.ndarray,
    query_qs: list[dict],
    target_embs: np.ndarray,
    target_qs: list[dict],
    api_key: str,
    target_name: str = "ターゲット",
    auto_threshold: float = AUTO_MATCH_THRESHOLD,
    llm_threshold: float = LLM_JUDGE_THRESHOLD,
) -> tuple[list[int], list[int], list[dict]]:
    """
    query質問群とtarget質問群のマッチングを実行。

    Returns:
        matched_indices: queryのうちtargetと類似するもののインデックス
        unmatched_indices: queryのうちtargetと類似しないもののインデックス
        match_details: 各queryの最良マッチ情報
    """
    print(f"\n  コサイン類似度計算中...")
    # 全query × 全target の類似度行列
    sim_matrix = cosine_similarity(query_embs, target_embs)

    # 各queryの最良マッチを取得
    best_indices = np.argmax(sim_matrix, axis=1)
    best_sims = np.max(sim_matrix, axis=1)

    # 3段階に分類
    auto_matched = []   # 自動マッチ（高類似度）
    llm_candidates = [] # LLM判定対象（微妙なライン）
    auto_unmatched = [] # 自動非マッチ（低類似度）

    for i in range(len(query_qs)):
        sim = best_sims[i]
        tidx = best_indices[i]
        info = {
            "query_idx": i,
            "target_idx": int(tidx),
            "similarity": float(sim),
            "target_question": target_qs[int(tidx)]["question"],
            "target_category": target_qs[int(tidx)]["category"],
        }
        if sim >= auto_threshold:
            auto_matched.append((i, info))
        elif sim >= llm_threshold:
            llm_candidates.append((i, info))
        else:
            auto_unmatched.append((i, info))

    print(f"  自動マッチ（≥{auto_threshold}）: {len(auto_matched)}件")
    print(f"  LLM判定対象（{llm_threshold}〜{auto_threshold}）: {len(llm_candidates)}件")
    print(f"  自動非マッチ（<{llm_threshold}）: {len(auto_unmatched)}件")

    # LLM判定
    llm_matched_indices = set()
    if llm_candidates:
        pairs_for_llm = [
            (
                query_qs[i]["question"],
                info["target_question"],
                info["similarity"],
            )
            for i, info in llm_candidates
        ]
        llm_results = llm_judge_pairs(pairs_for_llm, api_key, target_name)

        n_same = sum(1 for v in llm_results.values() if v)
        n_diff = sum(1 for v in llm_results.values() if not v)
        print(f"  LLM結果: 同={n_same}件, 別={n_diff}件, 未判定={len(llm_candidates)-len(llm_results)}件")

        for pair_idx, (i, info) in enumerate(llm_candidates):
            if llm_results.get(pair_idx, False):
                llm_matched_indices.add(i)

    # 最終結果の集約
    matched = set(i for i, _ in auto_matched) | llm_matched_indices
    all_details = {}
    for i, info in auto_matched + llm_candidates + auto_unmatched:
        info["matched"] = i in matched
        info["method"] = (
            "auto" if i in set(idx for idx, _ in auto_matched)
            else "llm" if i in set(idx for idx, _ in llm_candidates)
            else "auto_no"
        )
        all_details[i] = info

    matched_indices = sorted(matched)
    unmatched_indices = sorted(set(range(len(query_qs))) - matched)

    print(f"  最終: マッチ{len(matched_indices)}件, 非マッチ{len(unmatched_indices)}件")
    return matched_indices, unmatched_indices, all_details


# ============================================================
# カテゴリ誤分類の修正
# ============================================================
# サブカテゴリ名→正しいカテゴリのマッピング（キーワードベース）
_SUBCATEGORY_REMAP = {
    # 交通カテゴリに誤分類されやすいサブカテゴリ
    "ごみ処理": "環境・ごみ",
    "ごみのポイ捨て・路上喫煙": "環境・ごみ",
    "事業系ごみ・産業廃棄物": "環境・ごみ",
    "し尿くみ取り": "環境・ごみ",
    "公害": "環境・ごみ",
    "騒音等": "環境・ごみ",
    "アイドリング・ストップ": "環境・ごみ",
    "地球温暖化対策": "環境・ごみ",
    "水道料金": "住まい・生活",
    "上水道": "住まい・生活",
    "下水道": "住まい・生活",
    "河川": "住まい・生活",
    "住民組織": "市政・行政",
    "届け出・問い合わせ（上水道）": "住まい・生活",
    "浄水場": "住まい・生活",
}

# 質問文キーワード→正しいカテゴリ（サブカテゴリで判定できない場合のフォールバック）
_QUESTION_KEYWORD_REMAP = [
    (["蛍光灯", "ごみ", "リサイクル", "分別", "廃棄"], "環境・ごみ"),
    (["水道", "下水道", "浄水", "漏水"], "住まい・生活"),
    (["自治会", "町内会"], "市政・行政"),
    (["汚水", "養豚", "養鶏"], "環境・ごみ"),
    (["公園", "球技場", "テニスコート", "お花見"], "施設"),
    (["街灯", "防犯灯"], "住まい・生活"),
    (["環境基本条例", "環境に関する講座"], "環境・ごみ"),
]


def fix_category(question: str, current_category: str, subcategory: str) -> str:
    """誤分類されたカテゴリを修正する。

    サブカテゴリ名とキーワードから正しいカテゴリを推定。
    修正が不要な場合は元のカテゴリをそのまま返す。
    """
    # サブカテゴリ名で判定
    sub_clean = subcategory.strip()
    if sub_clean in _SUBCATEGORY_REMAP:
        new_cat = _SUBCATEGORY_REMAP[sub_clean]
        if new_cat != current_category:
            return new_cat

    # キーワードで判定
    for keywords, correct_cat in _QUESTION_KEYWORD_REMAP:
        if any(kw in question for kw in keywords) and correct_cat != current_category:
            return correct_cat

    return current_category


# ============================================================
# 不適切な提案質問のフィルタリング
# ============================================================
def filter_unsuitable_proposals(
    proposals: list[tuple[int, dict, dict]],
    city_name: str = "",
) -> tuple[list[tuple[int, dict, dict]], int]:
    """対象自治体のFAQとして不適切な質問を除外する。

    除外基準:
      1. 大都市固有の施設・サービス（市営地下鉄、中央卸売市場、区役所等）
      2. 時限的な制度・過去の施策（特定年度の給付金、終了イベント等）
      3. 内容が古い・時代遅れ（旧制度の説明等）
      4. 他自治体固有の制度名・施設名が質問文に残っているもの

    Returns:
        (filtered_proposals, n_removed)
    """
    filtered = []
    n_removed = 0

    for idx, q, detail in proposals:
        question = q["question"]
        remove = False
        reason = ""

        # --- 1. 大都市固有の施設・サービス ---
        big_city_services = [
            ("市バス・地下鉄", "市営地下鉄・市営バス"),
            ("市バス・市営地下鉄", "市営地下鉄・市営バス"),
            ("市営地下鉄", "市営地下鉄"),
            ("新交通について教えて", "新交通システム"),
            ("フェリー乗り場はどこ", "フェリー航路"),
            ("中央卸売市場の見学", "中央卸売市場"),
            ("区以外の区役所で", "区制度"),
        ]
        for kw, svc in big_city_services:
            if kw in question:
                remove, reason = True, f"大都市固有: {svc}"
                break

        # --- 2. 大都市・特定地域のみの制度 ---
        if not remove:
            big_city_systems = [
                ("事業所税", "事業所税は人口30万以上の都市のみ"),
                ("道民税", "北海道のみ"),
                ("都民税", "東京都のみ"),
                ("古都保存区域", "古都のみ"),
            ]
            for kw, reason_text in big_city_systems:
                if kw in question:
                    remove, reason = True, f"特定地域のみ: {reason_text}"
                    break

        # --- 3. 時限的な制度・過去の施策 ---
        if not remove:
            import re
            # 「令和○年度の制度変更」「202X年X月までに実施した」等
            if re.search(
                r'(令和[０-９\d]+年度の制度変更|'
                r'20\d{2}年\d+月まで(に実施|で終了)|'
                r'【20\d{2}年度分】|【令和[０-９\d]+年度分】)',
                question
            ):
                remove, reason = True, "時限的な施策"

            # 廃止済みの制度
            if not remove and "廃止しました" in question:
                remove, reason = True, "廃止済みの制度"

            # 特定イベント（万博、オリンピック、ワールドカップ等）
            if not remove and re.search(
                r'(万博|オリンピック|ワールドカップ|ワールドマスターズ)'
                r'.*(チケット|入場|エントリー|申込)',
                question
            ):
                remove, reason = True, "特定イベント"

        # --- 4. 内容が古い・時代遅れ ---
        if not remove:
            outdated = [
                ("外国人登録原票には載っていた", "旧外国人登録制度"),
                ("平成24年7月9日の法改正で、外国人", "2012年の制度変更"),
            ]
            for kw, reason_text in outdated:
                if kw in question:
                    remove, reason = True, f"時代遅れ: {reason_text}"
                    break

        # --- 5. 他自治体固有の制度名・組織名が残っている ---
        if not remove:
            # 特定自治体の固有制度・施設・組織名
            specific_names = [
                "ＩＣＯＣＡ", "ICOCA",
                "計量検査係",
                "ナビダイヤルを導入",
                "暮らしのガイドブック",
                "行政放送の過去の映像",
                "債権者登録番号",
            ]
            for name in specific_names:
                if name in question:
                    remove, reason = True, f"他自治体固有: {name}"
                    break

        if remove:
            n_removed += 1
            print(f"    除外: {reason} → {question[:55]}")
        else:
            filtered.append((idx, q, detail))

    return filtered, n_removed


# ============================================================
# Excel出力
# ============================================================
def save_results(
    proposals: list[tuple[int, dict, dict]],
    soumu_matched: list[tuple[int, dict, dict]],
    kuwana_matched: list[tuple[int, dict, dict]],
    output_path: str,
    query_qs: list[dict],
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    hf = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    def write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    def write_row(ws, ri, vals):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.alignment, c.border = cf, ca, tb

    # --- サマリー ---
    ws0 = wb.active
    ws0.title = "サマリー"
    summary = [
        ("処理概要", ""),
        ("代表質問（入力）", len(query_qs)),
        ("総務省FAQと重複（除外）", len(soumu_matched)),
        ("対象自治体FAQに既存（除外）", len(kuwana_matched)),
        ("提案FAQ（出力）", len(proposals)),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws0.cell(row=ri, column=1, value=k).font = Font(
            name="Arial", size=11, bold=(ri <= 1)
        )
        ws0.cell(row=ri, column=2, value=v).font = cf
    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 10

    from collections import Counter
    cat_counts = Counter(q["category"] for _, q, _ in proposals)
    ri = len(summary) + 2
    ws0.cell(row=ri, column=1, value="カテゴリ別提案件数").font = Font(
        name="Arial", size=11, bold=True
    )
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        ri += 1
        ws0.cell(row=ri, column=1, value=cat).font = cf
        ws0.cell(row=ri, column=2, value=cnt).font = cf

    # --- 提案FAQ ---
    ws1 = wb.create_sheet("提案FAQ")
    headers1 = [
        "No.", "カテゴリ", "サブカテゴリ", "提案質問", "参考回答",
        "出典自治体", "対象市FAQ最類似", "類似度",
    ]
    write_header(ws1, headers1)
    for ri, (idx, q, detail) in enumerate(proposals, 2):
        write_row(ws1, ri, [
            ri - 1, q["category"], q["subcategory"],
            q["question"], q["answer"][:400], q.get("source", ""),
            detail.get("target_question", "")[:80],
            f"{detail.get('similarity', 0):.3f}",
        ])
    for col, w in [("A", 6), ("B", 18), ("C", 18), ("D", 55),
                    ("E", 65), ("F", 12), ("G", 45), ("H", 8)]:
        ws1.column_dimensions[col].width = w
    ws1.auto_filter.ref = f"A1:H{len(proposals) + 1}"
    ws1.freeze_panes = "A2"

    # --- 総務省重複 ---
    ws2 = wb.create_sheet("総務省FAQ重複（除外）")
    headers2 = [
        "No.", "カテゴリ", "代表質問", "総務省FAQ類似質問",
        "類似度", "判定方法",
    ]
    write_header(ws2, headers2)
    for ri, (idx, q, detail) in enumerate(soumu_matched, 2):
        write_row(ws2, ri, [
            ri - 1, q["category"], q["question"][:70],
            detail.get("target_question", "")[:70],
            f"{detail.get('similarity', 0):.3f}",
            detail.get("method", ""),
        ])
    for col, w in [("A", 6), ("B", 18), ("C", 55), ("D", 55),
                    ("E", 8), ("F", 10)]:
        ws2.column_dimensions[col].width = w

    # --- 対象市既存 ---
    ws3 = wb.create_sheet("対象市FAQ既存（除外）")
    headers3 = [
        "No.", "カテゴリ", "代表質問", "対象市FAQ類似質問",
        "類似度", "判定方法",
    ]
    write_header(ws3, headers3)
    for ri, (idx, q, detail) in enumerate(kuwana_matched, 2):
        write_row(ws3, ri, [
            ri - 1, q["category"], q["question"][:70],
            detail.get("target_question", "")[:70],
            f"{detail.get('similarity', 0):.3f}",
            detail.get("method", ""),
        ])
    for col, w in [("A", 6), ("B", 18), ("C", 55), ("D", 55),
                    ("E", 8), ("F", 10)]:
        ws3.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n出力: {output_path}")


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="FAQ ギャップ分析（Embedding + LLM判定）"
    )
    parser.add_argument(
        "--daihyo", required=True, help="代表質問Excelファイル"
    )
    parser.add_argument(
        "--soumu", required=True, help="総務省FAQのExcelファイル"
    )
    parser.add_argument(
        "--city", required=True, help="対象自治体FAQのExcelファイル"
    )
    parser.add_argument(
        "--output", "-o", default="./FAQ提案.xlsx", help="出力Excelファイル"
    )
    parser.add_argument(
        "--api-key", help="Gemini APIキー（環境変数 GEMINI_API_KEY でも可）"
    )
    parser.add_argument(
        "--auto-threshold", type=float, default=AUTO_MATCH_THRESHOLD,
        help=f"自動マッチ閾値 (default: {AUTO_MATCH_THRESHOLD})"
    )
    parser.add_argument(
        "--llm-threshold", type=float, default=LLM_JUDGE_THRESHOLD,
        help=f"LLM判定閾値 (default: {LLM_JUDGE_THRESHOLD})"
    )
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: Gemini APIキーが必要です。")
        print("  --api-key YOUR_KEY または GEMINI_API_KEY=YOUR_KEY")
        sys.exit(1)

    print("=" * 60)
    print("FAQ ギャップ分析（Embedding + LLM判定）")
    print(f"  代表質問: {args.daihyo}")
    print(f"  総務省FAQ: {args.soumu}")
    print(f"  対象自治体FAQ: {args.city}")
    print(f"  出力: {args.output}")
    print(f"  閾値: 自動マッチ≥{args.auto_threshold}, "
          f"LLM判定≥{args.llm_threshold}")
    print("=" * 60)

    # ① 読み込み
    print("\n【STEP 1】データ読み込み")
    daihyo_qs = load_questions(args.daihyo)
    soumu_qs = load_questions(args.soumu, skip_summary=False)
    city_qs = load_questions(args.city)
    print(f"  代表質問: {len(daihyo_qs)}件")
    print(f"  総務省FAQ: {len(soumu_qs)}件")
    print(f"  対象自治体FAQ: {len(city_qs)}件")

    # ② Embedding
    print("\n【STEP 2】Embedding（Gemini API）")
    all_texts = (
        [q["question"] for q in daihyo_qs]
        + [q["question"] for q in soumu_qs]
        + [q["question"] for q in city_qs]
    )
    all_embs = embed_with_gemini(all_texts, api_key)

    n_d = len(daihyo_qs)
    n_s = len(soumu_qs)
    n_c = len(city_qs)
    emb_daihyo = all_embs[:n_d]
    emb_soumu = all_embs[n_d:n_d + n_s]
    emb_city = all_embs[n_d + n_s:]

    # ③ 代表質問 vs 総務省FAQ
    print("\n【STEP 3】代表質問 vs 総務省FAQ")
    soumu_matched_idx, soumu_unmatched_idx, soumu_details = match_questions(
        emb_daihyo, daihyo_qs, emb_soumu, soumu_qs, api_key,
        target_name="総務省FAQ",
        auto_threshold=args.auto_threshold,
        llm_threshold=args.llm_threshold,
    )

    # ④ 残り vs 対象自治体FAQ
    print("\n【STEP 4】残り vs 対象自治体FAQ")
    # 総務省FAQと非マッチだったもののみを対象にする
    remaining_embs = emb_daihyo[soumu_unmatched_idx]
    remaining_qs = [daihyo_qs[i] for i in soumu_unmatched_idx]
    remaining_orig_idx = soumu_unmatched_idx  # 元のインデックスを保持

    city_matched_idx, city_unmatched_idx, city_details = match_questions(
        remaining_embs, remaining_qs, emb_city, city_qs, api_key,
        target_name="対象自治体FAQ",
        auto_threshold=args.auto_threshold,
        llm_threshold=args.llm_threshold,
    )

    # ⑤ 結果整理
    print("\n【STEP 5】結果整理")

    # 総務省重複
    soumu_matched_list = [
        (i, daihyo_qs[i], soumu_details[i])
        for i in soumu_matched_idx
    ]

    # 対象市既存
    kuwana_matched_list = [
        (
            remaining_orig_idx[i],
            remaining_qs[i],
            city_details[i],
        )
        for i in city_matched_idx
    ]

    # 提案
    proposal_list = [
        (
            remaining_orig_idx[i],
            remaining_qs[i],
            city_details[i],
        )
        for i in city_unmatched_idx
    ]

    print(f"  代表質問: {len(daihyo_qs)}件")
    print(f"  → 総務省FAQと重複: {len(soumu_matched_list)}件")
    print(f"  → 対象自治体FAQに既存: {len(kuwana_matched_list)}件")
    print(f"  → 提案候補: {len(proposal_list)}件")

    # ⑥ カテゴリ誤分類の修正
    print("\n【STEP 6】カテゴリ誤分類の修正")
    n_fixed = 0
    for i, (idx, q, detail) in enumerate(proposal_list):
        old_cat = q["category"]
        new_cat = fix_category(q["question"], old_cat, q["subcategory"])
        if new_cat != old_cat:
            q["category"] = new_cat
            n_fixed += 1
            print(f"  修正: [{old_cat}]→[{new_cat}] {q['question'][:50]}")
    print(f"  カテゴリ修正: {n_fixed}件")

    # ⑦ 不適切な提案の除外
    print("\n【STEP 7】不適切な提案質問の除外")
    proposal_list, n_unsuitable = filter_unsuitable_proposals(proposal_list)
    print(f"  除外: {n_unsuitable}件 → 残り{len(proposal_list)}件")

    # ⑧ Excel出力
    print("\n【STEP 8】Excel出力")
    save_results(
        proposal_list, soumu_matched_list, kuwana_matched_list,
        args.output, daihyo_qs,
    )

    print("\n" + "=" * 60)
    print("処理完了！")
    print(f"  提案FAQ: {len(proposal_list)}件")
    print(f"  カテゴリ修正: {n_fixed}件")
    print(f"  不適切除外: {n_unsuitable}件")
    print(f"  出力: {args.output}")
    print("=" * 60)


if __name__ == "__main__":
    main()