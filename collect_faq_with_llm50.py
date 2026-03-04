# -*- coding: utf-8 -*-
"""
LLMガイド付きFAQ収集スクリプト（改良版）
サイト分析結果を元に、Gemini API + Google Search Grounding でQ&Aを効率的に収集する。

改善点:
- バッチ方式（4分割）でAPI呼び出し → トークン切れ防止、各カテゴリの質が向上
- 解析結果（カテゴリ・サブカテゴリ・サンプルFAQ）をバッチプロンプトに埋め込み
- Google Search Grounding を全リトライで維持し、プロンプトでも検索を明示指示
- 件数不足時の自動補完バッチ
"""

import json
import os
import re
import time
import argparse
from collections import Counter, defaultdict
from pathlib import Path

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================
# 設定
# ============================================================
_BASE_DIR = Path(__file__).parent
ANALYSIS_DIR = str(_BASE_DIR / "analysis_results")
OUTPUT_DIR = str(_BASE_DIR / "faq_data2")
MUNICIPALITIES_FILE = str(_BASE_DIR / "municipalities.json")
GEMINI_API_KEY = "geminiのAPIキーを入力してください"

TARGET_PER_CITY = 50          # 自治体あたり目標件数
NUM_BATCHES = 2                # バッチ数
NUM_CATEGORIES_PER_BATCH = 8   # バッチあたりカテゴリ数
MIN_PER_BATCH = max(TARGET_PER_CITY // NUM_BATCHES, 5)          # 自動計算
MIN_PER_CATEGORY = max(TARGET_PER_CITY // (NUM_BATCHES * NUM_CATEGORIES_PER_BATCH), 1)  # 自動計算
BATCH_WAIT_SEC = 5             # バッチ間の待機秒数
CITY_WAIT_SEC = 10             # 自治体間の待機秒数

# ============================================================
# カテゴリバッチ定義（16カテゴリを2グループに分割）
# ============================================================
CATEGORY_BATCHES = [
    {
        "name": "住民・届出・税金・保険・福祉・医療",
        "categories": [
            ("住民票・戸籍・印鑑登録", "転入届、転出届、住民票の写し、戸籍謄本・抄本、印鑑登録、マイナンバーカード、婚姻届、離婚届、出生届、死亡届"),
            ("届出・証明", "各種証明書の取得方法、届出方法、コンビニ交付、郵送請求"),
            ("税金", "住民税、固定資産税、軽自動車税、納税証明書、確定申告、税金の納付方法、口座振替、市税の減免"),
            ("年金", "国民年金の加入、免除・猶予、受給、障害年金、遺族年金"),
            ("国民健康保険", "加入・脱退、保険料の計算・納付、給付、高額療養費、限度額適用認定証、人間ドック助成"),
            ("介護保険", "認定申請、保険料、在宅サービス、施設サービス、福祉用具、住宅改修"),
            ("福祉", "障害者福祉、障害者手帳、高齢者福祉、生活保護、民生委員"),
            ("健康・医療", "予防接種、がん検診、健康診断、妊婦健診、乳幼児健診、休日・夜間診療"),
        ],
    },
    {
        "name": "子育て・教育・暮らし・安全・環境・行政",
        "categories": [
            ("子育て・教育", "保育所・保育園の入所、児童手当、児童扶養手当、幼稚園、小中学校の入学・転校、就学援助、放課後児童クラブ、子ども医療費助成"),
            ("ごみ・リサイクル", "家庭ごみの分別方法、収集日・収集場所、粗大ごみの出し方、持ち込み処分、資源ごみ、ごみ袋"),
            ("上下水道", "水道の開栓・閉栓、水道料金、漏水、下水道の接続、下水道使用料"),
            ("住宅・土地", "市営住宅の申込、建築確認申請、都市計画、用途地域"),
            ("防災・安全", "避難所、防災マップ、ハザードマップ、AED設置場所、消防、救急"),
            ("道路・交通", "道路の損傷・補修要望、交通安全、駐輪場・駐車場、街灯"),
            ("環境", "騒音・振動、犬・猫、害虫・害獣、空き地の雑草"),
            ("市政・行政", "情報公開請求、パブリックコメント、選挙、市議会、広報紙、市の施設"),
        ],
    },
]


# ============================================================
# Gemini API 初期化
# ============================================================
def setup_gemini(api_key):
    """Gemini APIの初期化"""
    client = genai.Client(api_key=api_key)
    return client


# ============================================================
# 解析結果からバッチに関連するカテゴリ情報を抽出
# ============================================================
def extract_relevant_analysis(analysis_data, batch):
    """
    analyze_faq_site.py の解析結果から、
    当該バッチのカテゴリに関連するカテゴリ・サブカテゴリ・FAQリンクを抽出する。
    """
    if not analysis_data:
        return "", "", ""

    batch_keywords = []
    for cat_name, keywords_str in batch["categories"]:
        batch_keywords.append(cat_name)
        batch_keywords.extend(k.strip() for k in keywords_str.split("、"))

    # --- サイトカテゴリ情報 ---
    site_categories_text = ""
    if analysis_data.get("categories"):
        for cat in analysis_data["categories"]:
            cat_name_lower = cat["name"].lower()
            # バッチのキーワードとカテゴリ名を照合（部分一致）
            is_relevant = any(
                kw in cat_name_lower or cat_name_lower in kw
                for kw in batch_keywords
            )
            if is_relevant:
                site_categories_text += f"\n  ■ {cat['name']}\n"
                for sub in cat.get("subcategories", []):
                    site_categories_text += f"    - {sub['name']}: {sub['url']}\n"
                for faq in cat.get("faq_links", []):
                    site_categories_text += f"    - Q: {faq.get('title', '')}\n"

        # 関連カテゴリが見つからなければ全カテゴリを渡す（情報がないよりまし）
        if not site_categories_text:
            for cat in analysis_data["categories"]:
                site_categories_text += f"\n  ■ {cat['name']}\n"
                for sub in cat.get("subcategories", [])[:5]:
                    site_categories_text += f"    - {sub['name']}: {sub['url']}\n"

    # --- 直接FAQリンク ---
    direct_links_text = ""
    if analysis_data.get("direct_faq_links"):
        relevant_links = []
        for link in analysis_data["direct_faq_links"]:
            title = link.get("title", "")
            if any(kw in title for kw in batch_keywords):
                relevant_links.append(link)
        # 関連リンクが少なければ全体から補完
        if len(relevant_links) < 5:
            relevant_links = analysis_data["direct_faq_links"][:15]
        for link in relevant_links[:15]:
            direct_links_text += f"  - {link['title']} ({link.get('url', '')})\n"

    # --- サンプルFAQ ---
    sample_text = ""
    if analysis_data.get("sample_faqs"):
        for s in analysis_data["sample_faqs"]:
            sample_text += f"\n  質問: {s.get('question', '不明')}\n"
            sample_text += f"  回答（一部）: {s.get('answer_preview', '不明')[:200]}\n"

    return site_categories_text, direct_links_text, sample_text


# ============================================================
# プロンプト生成
# ============================================================
def build_batch_prompt(city_name, batch, analysis_data=None):
    """
    バッチ（カテゴリグループ）ごとのプロンプトを生成。
    解析結果があれば、関連するサイト構造情報を埋め込む。
    Google Search を使って実際のサイトを検索するよう明示的に指示する。
    """
    faq_url = ""
    if analysis_data:
        faq_url = analysis_data.get("actual_url", analysis_data.get("faq_url", ""))

    # バッチ内カテゴリの詳細
    categories_detail = ""
    for cat_name, keywords in batch["categories"]:
        categories_detail += f"\n### {cat_name}\n対象トピック例: {keywords}\n"

    # 解析結果から関連情報を抽出
    site_categories_text, direct_links_text, sample_text = extract_relevant_analysis(
        analysis_data, batch
    )

    # 解析結果セクション（データがある場合のみ挿入）
    analysis_section = ""
    if site_categories_text or direct_links_text or sample_text:
        analysis_section = f"""
## サイト構造の分析結果（analyze_faq_site.py による自動解析）
以下は {city_name} のFAQサイトをスクレイピングして得られた情報です。
Google検索でこのサイトにアクセスし、以下のカテゴリ・URLを手がかりにFAQを収集してください。

### サイトのカテゴリ構造
{site_categories_text if site_categories_text else "（自動解析では検出されず）"}

### サイトで発見されたFAQ質問
{direct_links_text if direct_links_text else "（自動解析では検出されず）"}

### サンプルFAQの内容（回答の書き方の参考）
{sample_text if sample_text else "（サンプル取得なし）"}
"""

    prompt = f"""あなたは自治体のFAQデータを網羅的に収集する専門家です。
Google検索を使って、対象自治体のFAQサイトから実際の質問と回答を収集してください。

## 対象自治体: {city_name}
## FAQサイトURL: {faq_url}
## 担当カテゴリグループ: {batch['name']}
{analysis_section}
## 収集対象カテゴリ（以下の4カテゴリに集中してください）
{categories_detail}

## Google検索の使い方
以下のように検索して、{city_name}のFAQサイトから実際の質問と回答を見つけてください:
- 「{city_name} FAQ [カテゴリ名]」で検索
- 「{city_name} よくある質問 [トピック名]」で検索
- FAQサイトURL「{faq_url}」内のページを検索
- 各カテゴリについて複数回検索し、できるだけ多くの質問を収集してください

## 出力形式
JSON配列のみを出力してください。説明文やコメントは一切不要です。

```json
[
  {{
    "category": "カテゴリ名",
    "subcategory": "サブカテゴリ名",
    "question": "質問文（疑問形で）",
    "answer": "回答文（具体的に。必要書類、届出場所、費用、期限等を含む）"
  }}
]
```

## 重要な注意事項
- **Google検索を必ず実行**: 検索せずに知識だけで回答しないでください。必ずGoogle検索で {city_name} のFAQサイトを参照してください
- **各カテゴリ最低{MIN_PER_CATEGORY}件以上**: {NUM_CATEGORIES_PER_BATCH}カテゴリそれぞれから{MIN_PER_CATEGORY}件以上、合計{MIN_PER_BATCH}件以上を出力してください
- **回答は具体的に**: {city_name}の窓口名、電話番号、受付時間、費用、必要書類など固有情報をできるだけ含めてください
- **JSON形式厳守**: 出力はJSON配列のみ。マークダウンのコードブロック（```json ...```）で囲んでください
"""
    return prompt


def build_supplement_prompt(city_name, weak_categories, analysis_data=None):
    """
    件数不足カテゴリの補完用プロンプトを生成。
    """
    faq_url = ""
    if analysis_data:
        faq_url = analysis_data.get("actual_url", analysis_data.get("faq_url", ""))

    categories_list = ""
    for cat_name, current_count in weak_categories:
        need = max(MIN_PER_CATEGORY - current_count, 5)
        categories_list += f"- **{cat_name}**: 現在{current_count}件 → あと{need}件以上追加してください\n"

    prompt = f"""あなたは自治体のFAQデータを網羅的に収集する専門家です。
Google検索を使って、不足している質問を追加収集してください。

## 対象自治体: {city_name}
## FAQサイトURL: {faq_url}

## 補完が必要なカテゴリ
以下のカテゴリで質問数が不足しています。Google検索で「{city_name} FAQ [カテゴリ名]」「{city_name} よくある質問 [トピック名]」などを検索し、追加の質問と回答を収集してください。

{categories_list}

## 出力形式
JSON配列のみを出力してください。

```json
[
  {{
    "category": "カテゴリ名",
    "subcategory": "サブカテゴリ名",
    "question": "質問文（疑問形で）",
    "answer": "回答文（具体的に。{city_name}固有の情報を含む）"
  }}
]
```

## 重要
- Google検索を必ず実行して {city_name} の公式情報を参照してください
- 既に収集済みの質問と重複しない新しい質問を出してください
- JSON形式厳守
"""
    return prompt


# ============================================================
# Gemini API 呼び出し（Google Search Grounding 対応）
# ============================================================
def call_gemini_with_search(client, prompt, use_grounding=True, max_retries=4):
    """
    Gemini APIを呼び出す。
    Google Search Grounding を有効化し、全リトライで維持する。
    """


    for attempt in range(max_retries):
        try:
            config = types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=65536,
            )
            if use_grounding:
                config.tools = [types.Tool(google_search=types.GoogleSearch())]

            response = client.models.generate_content(
                model="gemini-3-flash-preview",
                contents=prompt,
                config=config,
            )

            return response.text

        except Exception as e:
            error_str = str(e)

            # レート制限 → 段階的待機
            if "429" in error_str or "quota" in error_str.lower() or "rate" in error_str.lower():
                wait = 30 * (attempt + 1)
                print(f"    レート制限検出。{wait}秒待機後リトライ... (attempt {attempt+1}/{max_retries})")
                time.sleep(wait)
                continue

            # Grounding固有エラー → Groundingなしにフォールバック
            if use_grounding and ("grounding" in error_str.lower() or "tool" in error_str.lower()):
                print(f"    Groundingエラー: {e}")
                print(f"    Groundingなしでリトライ...")
                use_grounding = False
                continue

            # その他のエラー
            print(f"    APIエラー (attempt {attempt+1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(10)
                continue
            else:
                return None

    print(f"    最大リトライ回数に到達")
    return None


# ============================================================
# JSON パース（修復機能付き）
# ============================================================
def parse_llm_response(response_text):
    """LLMの応答からJSONデータを抽出する（修復機能付き）"""
    # コードブロック内のJSONを抽出
    json_match = re.search(r"```json\s*\n?(.*?)\n?\s*```", response_text, re.DOTALL)
    if json_match:
        json_str = json_match.group(1).strip()
    else:
        start = response_text.find("[")
        if start >= 0:
            json_str = response_text[start:].strip()
        else:
            json_str = response_text.strip()

    # そのまま解析
    try:
        data = json.loads(json_str)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # 不完全なJSONの修復
    print("    JSON修復を試行中...")
    last_complete = -1
    brace_depth = 0
    in_string = False
    escape_next = False

    for i, ch in enumerate(json_str):
        if escape_next:
            escape_next = False
            continue
        if ch == '\\':
            escape_next = True
            continue
        if ch == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == '{':
            brace_depth += 1
        elif ch == '}':
            brace_depth -= 1
            if brace_depth == 0:
                last_complete = i

    if last_complete > 0:
        repaired = json_str[:last_complete + 1] + "\n]"
        try:
            data = json.loads(repaired)
            if isinstance(data, list):
                print(f"    修復成功: {len(data)}件")
                return data
        except json.JSONDecodeError as e:
            print(f"    修復後も解析エラー: {e}")

    # フォールバック: 個別オブジェクト抽出
    print("    個別抽出を試行中...")
    items = []
    for m in re.finditer(r'\{[^{}]*"category"[^{}]*"question"[^{}]*\}', json_str):
        try:
            item = json.loads(m.group())
            items.append(item)
        except Exception:
            pass
    if items:
        print(f"    個別抽出成功: {len(items)}件")
        return items

    return []


# ============================================================
# Excel 保存
# ============================================================
def save_to_excel(city_name, faq_data):
    """収集したFAQデータをExcelファイルに保存"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_path = os.path.join(OUTPUT_DIR, f"{city_name}FAQ.xlsx")

    wb = openpyxl.Workbook()

    # スタイル定義
    hf = Font(name="Yu Gothic", size=12, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    nf = Font(name="Yu Gothic", size=10)
    na = Alignment(vertical="top", wrap_text=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # カテゴリでグループ化
    cat_groups = defaultdict(list)
    for item in faq_data:
        cat = item.get("category", "その他")
        cat_groups[cat].append(item)

    for cat_name, items in cat_groups.items():
        sheet_name = cat_name.replace("/", "・").replace("\\", "・")[:28] + " FAQ"
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # ヘッダー
        headers = ["No.", "サブカテゴリ", "質問", "回答"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

        # データ
        for rn, item in enumerate(items, 2):
            fn = rn - 1
            ws.cell(row=rn, column=1, value=fn).font = nf
            ws.cell(row=rn, column=1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=rn, column=1).border = tb

            for col, key in [(2, "subcategory"), (3, "question"), (4, "answer")]:
                val = item.get(key, "")
                c = ws.cell(row=rn, column=col, value=val)
                c.font, c.alignment, c.border = nf, na, tb

        # 列幅
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 80
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:D{len(items) + 1}"

    # デフォルトの空シートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(excel_path)
    return excel_path


# ============================================================
# メイン収集ロジック（バッチ方式）
# ============================================================
def collect_city(model, city_info, use_grounding=True):
    """
    1つの自治体のFAQをバッチ方式で収集する。
    1. 解析結果を読み込み
    2. 4バッチに分けてAPI呼び出し（各バッチに関連する解析情報を埋め込み）
    3. 件数不足カテゴリがあれば補完バッチを実行
    4. 全結果を統合してExcel保存
    """
    city_name = city_info["name"]
    faq_url = city_info["faq_url"]

    print(f"\n{'='*60}")
    print(f"■ 収集中: {city_name}")
    print(f"  FAQ URL: {faq_url}")
    print(f"{'='*60}")

    # --- 解析結果の読み込み ---
    analysis_path = os.path.join(ANALYSIS_DIR, f"{city_name}.json")
    analysis_data = None
    if os.path.exists(analysis_path):
        with open(analysis_path, "r", encoding="utf-8") as f:
            analysis_data = json.load(f)
        cat_count = len(analysis_data.get("categories", []))
        faq_count = analysis_data.get("total_direct_faq_links", 0)
        print(f"  ✅ 解析データ読み込み済み（サイトカテゴリ: {cat_count}件, FAQリンク: {faq_count}件）")
    else:
        print(f"  ⚠ 解析データなし → 定型カテゴリ + Google検索で収集します")

    # --- 4バッチでAPI呼び出し ---
    all_faq_data = []

    for batch_idx, batch in enumerate(CATEGORY_BATCHES, 1):
        print(f"\n  --- バッチ {batch_idx}/4: {batch['name']} ---")

        prompt = build_batch_prompt(city_name, batch, analysis_data)

        response_text = call_gemini_with_search(
            model, prompt, use_grounding
        )

        if not response_text:
            print(f"    ⚠ バッチ {batch_idx} レスポンス取得失敗。スキップ。")
            continue

        print(f"    応答文字数: {len(response_text)}")

        # JSON パース
        batch_data = parse_llm_response(response_text)
        print(f"    解析結果: {len(batch_data)}件")

        if not batch_data:
            # デバッグ用に生レスポンスを保存
            debug_path = os.path.join(OUTPUT_DIR, f"{city_name}_batch{batch_idx}_raw.txt")
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(response_text)
            print(f"    ⚠ 解析失敗。生レスポンス保存: {debug_path}")
            continue

        all_faq_data.extend(batch_data)

        # カテゴリ別件数を表示
        batch_counts = Counter(item.get("category", "その他") for item in batch_data)
        for cat, cnt in batch_counts.most_common():
            print(f"      {cat}: {cnt}件")

        # バッチ間のレート制限対策
        if batch_idx < len(CATEGORY_BATCHES):
            print(f"    {BATCH_WAIT_SEC}秒待機...")
            time.sleep(BATCH_WAIT_SEC)

    # --- 件数チェック & 補完バッチ ---
    if all_faq_data:
        all_faq_data = _run_supplement_if_needed(
            model, city_name, all_faq_data, analysis_data, use_grounding
        )

    # --- 結果集計 ---
    print(f"\n  === {city_name} 収集完了 ===")
    print(f"  合計: {len(all_faq_data)}件")

    if not all_faq_data:
        print(f"  ⚠ データ0件。Excel保存をスキップします。")
        return [], None

    # カテゴリ別件数
    cat_counts = Counter(item.get("category", "その他") for item in all_faq_data)
    for cat, cnt in cat_counts.most_common():
        print(f"    {cat}: {cnt}件")

    # --- Excel保存 ---
    excel_path = save_to_excel(city_name, all_faq_data)
    print(f"  💾 保存完了: {excel_path}")

    # --- JSON保存（デバッグ・再利用用） ---
    json_path = os.path.join(OUTPUT_DIR, f"{city_name}FAQ.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_faq_data, f, ensure_ascii=False, indent=2)

    return all_faq_data, excel_path



def _run_supplement_if_needed(model, city_name, all_faq_data, analysis_data, use_grounding):
    """件数不足カテゴリがあれば補完バッチを1回実行する"""
    cat_counts = Counter(item.get("category", "その他") for item in all_faq_data)

    # 全16カテゴリの期待リスト
    all_expected_categories = []
    for batch in CATEGORY_BATCHES:
        for cat_name, _ in batch["categories"]:
            all_expected_categories.append(cat_name)

    weak_categories = []
    for cat_name in all_expected_categories:
        count = cat_counts.get(cat_name, 0)
        if count < MIN_PER_CATEGORY:
            weak_categories.append((cat_name, count))

    if not weak_categories:
        print(f"\n  ✅ 全カテゴリ {MIN_PER_CATEGORY}件以上。補完不要。")
        return all_faq_data

    print(f"\n  --- 補完バッチ: {len(weak_categories)}カテゴリが不足 ---")
    for cat, cnt in weak_categories:
        print(f"    {cat}: {cnt}件 (不足)")

    prompt = build_supplement_prompt(city_name, weak_categories, analysis_data)
    response_text = call_gemini_with_search(model, prompt, use_grounding)

    if response_text:
        supplement_data = parse_llm_response(response_text)
        if supplement_data:
            print(f"    補完結果: {len(supplement_data)}件追加")
            all_faq_data.extend(supplement_data)
        else:
            print(f"    ⚠ 補完データの解析失敗")
    else:
        print(f"    ⚠ 補完バッチのレスポンス取得失敗")

    return all_faq_data


# ============================================================
# main
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="LLMガイド付きFAQ収集（バッチ方式 + Google Search Grounding）"
    )
    parser.add_argument("--api-key", help="Gemini APIキー（環境変数 GEMINI_API_KEY でも可）")
    parser.add_argument("--city", help="特定の自治体名を指定")
    parser.add_argument("--all", action="store_true", help="全pending自治体を収集")
    parser.add_argument("--no-grounding", action="store_true", help="Google Search Groundingを無効化")
    parser.add_argument("--status", action="store_true", help="処理状況を表示")
    args = parser.parse_args()

    # APIキー
    api_key = GEMINI_API_KEY or args.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key and not args.status:
        print("エラー: Gemini APIキーが必要です")
        print("  --api-key YOUR_KEY または $env:GEMINI_API_KEY='YOUR_KEY'")
        return

    # municipalities.json を読み込み
    with open(MUNICIPALITIES_FILE, "r", encoding="utf-8") as f:
        municipalities = json.load(f)

    if args.status:
        done = [m for m in municipalities if m["status"] == "done"]
        pending = [m for m in municipalities if m["status"] == "pending"]
        print(f"取得済み: {len(done)}市")
        print(f"未処理: {len(pending)}市")
        print(f"合計: {len(municipalities)}市")
        print()
        for m in municipalities:
            has_excel = os.path.exists(os.path.join(OUTPUT_DIR, f"{m['name']}FAQ.xlsx"))
            has_analysis = os.path.exists(os.path.join(ANALYSIS_DIR, f"{m['name']}.json"))
            status = "✅ Excel済" if has_excel else ("📊 解析済" if has_analysis else "⬜ 未処理")
            print(f"  {status} | {m['name']} ({m.get('size', '?')}) [{m.get('region', '?')}]")
        return

    # Gemini初期化
    model = setup_gemini(api_key)
    use_grounding = not args.no_grounding

    if use_grounding:
        print("🔍 Google Search Grounding: 有効")
    else:
        print("⚠ Google Search Grounding: 無効")

    if args.city:
        city = next((m for m in municipalities if m["name"] == args.city), None)
        if not city:
            print(f"エラー: '{args.city}' が見つかりません")
            return
        collect_city(model, city, use_grounding)

    elif args.all:
        to_process = municipalities
        print(f"\n処理対象: {len(to_process)}市")

        results = {}
        for i, city in enumerate(to_process, 1):
            print(f"\n[{i}/{len(to_process)}]")
            faq_data, excel_path = collect_city(model, city, use_grounding)
            results[city["name"]] = len(faq_data)

            # 自治体間のレート制限対策
            if i < len(to_process):
                print(f"\n  次の自治体まで{CITY_WAIT_SEC}秒待機...")
                time.sleep(CITY_WAIT_SEC)

        # 集計
        print(f"\n{'='*60}")
        print("📊 収集結果サマリー")
        print(f"{'='*60}")
        total = 0
        for name, count in results.items():
            status = "✅" if count >= TARGET_PER_CITY * 0.8 else "⚠"
            print(f"  {status} {name}: {count}件")
            total += count
        print(f"\n  合計: {total}件（{len(results)}自治体）")
        print(f"  平均: {total // max(len(results), 1)}件/自治体")

    else:
        parser.print_help()


if __name__ == "__main__":
    main()