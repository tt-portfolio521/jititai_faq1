# -*- coding: utf-8 -*-
"""
京丹後市FAQ収集スクリプト（v5 最終版）

【HTML構造（デバッグで判明）】
  <article id="contents">
    <h1><span><span>質問タイトル</span></span></h1>
    <div id="contents-in">          ← ★ここが回答エリア
      <div class="faq-answer-area">
        <div>
          <p>回答テキスト</p>
          ...関連リンク等...
        </div>
      </div>
      <div>                         ← 問い合わせ先
        <dl>
          <dt>この記事に関するお問い合わせ先</dt>
          <dd>担当課名、電話番号等</dd>
        </dl>
      </div>
    </div>
  </article>

【抽出方法】
  1. 質問: 最後のh1のテキスト
  2. 回答: div#contents-in のテキストから問い合わせ先の前まで
  3. 問い合わせ先: div#contents-in 内のdl > dd

使い方:
  pip install requests beautifulsoup4 openpyxl
  python collect_kyotango_faq.py
"""

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import re
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

BASE_URL = "https://www.city.kyotango.lg.jp"
FAQ_TOP_URL = f"{BASE_URL}/top/faq/index.html"
OUTPUT_FILE = "京丹後市FAQ.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

REQUEST_DELAY = 1.0


def fetch_page(url, timeout=30):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        resp.encoding = resp.apparent_encoding or "utf-8"
        if resp.status_code == 200:
            return BeautifulSoup(resp.text, "html.parser")
        print(f"  [WARN] ステータス {resp.status_code}: {url}")
        return None
    except Exception as e:
        print(f"  [ERROR] {url} → {e}")
        return None


def get_category_pages(faq_top_url):
    soup = fetch_page(faq_top_url)
    if not soup:
        return []
    categories = []
    current_main = ""
    for tag in soup.find_all(["h2", "li"]):
        if tag.name == "h2":
            a = tag.find("a")
            if a:
                current_main = a.get_text(strip=True)
        elif tag.name == "li":
            a = tag.find("a", href=True)
            if a and "/faq/" in a["href"]:
                sub_url = urljoin(faq_top_url, a["href"])
                if sub_url.endswith("index.html"):
                    categories.append({
                        "main_category": current_main,
                        "sub_category": a.get_text(strip=True),
                        "url": sub_url,
                    })
    print(f"  カテゴリ: {len(categories)}件")
    return categories


def get_faq_links(category_url):
    soup = fetch_page(category_url)
    if not soup:
        return []
    links, seen = [], set()
    for a in soup.find_all("a", href=True):
        full_url = urljoin(category_url, a["href"])
        text = a.get_text(strip=True)
        if (full_url not in seen and "/faq/" in full_url
                and not full_url.endswith("index.html")
                and full_url.endswith(".html") and text and len(text) > 5):
            seen.add(full_url)
            links.append({"title": text, "url": full_url})
    return links


def extract_faq_detail(url):
    """
    div#contents-in を回答エリアとして取得する。
    問い合わせ先のdlタグの前までが回答テキスト。
    """
    soup = fetch_page(url)
    if not soup:
        return None

    result = {"url": url}

    # ---- 質問: 最後のh1 ----
    all_h1 = soup.find_all("h1")
    faq_h1 = all_h1[-1] if all_h1 else None
    result["question"] = faq_h1.get_text(strip=True) if faq_h1 else ""

    # ---- 回答: div#contents-in から取得 ----
    answer = ""
    contents_in = soup.find("div", id="contents-in")

    if contents_in:
        # contents-in の中から問い合わせ先(dl)を探す
        inquiry_dl = None
        for dl in contents_in.find_all("dl"):
            if "この記事に関するお問い合わせ先" in dl.get_text():
                inquiry_dl = dl
                break

        if inquiry_dl:
            # dlの親divを特定（問い合わせ先ブロック全体）
            inquiry_parent = inquiry_dl.parent

            # contents-in直下の子要素を順に走査し、
            # 問い合わせ先ブロックの前までのテキストを回答として収集
            answer_parts = []
            for child in contents_in.children:
                # 問い合わせ先ブロックに到達したら停止
                if child is inquiry_parent or child is inquiry_dl:
                    break
                if hasattr(child, 'get_text'):
                    text = child.get_text(separator="\n", strip=True)
                    if text:
                        # HTMLコメント的なテキストを除外
                        lines = []
                        for line in text.split("\n"):
                            line = line.strip()
                            if (line and
                                not line.startswith("「") or
                                not line.endswith("」")):
                                # 「関連画像」「メールリンク」等のCMSコメントを除外
                                if not re.match(r'^「.+」$', line) and \
                                   not line.startswith("ここから") and \
                                   not line.startswith("ここまで"):
                                    lines.append(line)
                        if lines:
                            answer_parts.append("\n".join(lines))
            answer = "\n".join(answer_parts).strip()
        else:
            # dlが見つからない場合: contents-in全体のテキスト
            full_text = contents_in.get_text(separator="\n", strip=True)
            # CMSコメントを除外
            lines = []
            for line in full_text.split("\n"):
                line = line.strip()
                if (line and
                    not re.match(r'^「.+」$', line) and
                    not line.startswith("ここから") and
                    not line.startswith("ここまで")):
                    lines.append(line)
            answer = "\n".join(lines).strip()

    # フォールバック: article#contentsから取得
    if not answer:
        article = soup.find("article", id="contents")
        if article:
            text = article.get_text(separator="\n", strip=True)
            q = result.get("question", "")
            if q and q in text:
                after_q = text.split(q, 1)[-1].strip()
                if "この記事に関するお問い合わせ先" in after_q:
                    after_q = after_q.split("この記事に関するお問い合わせ先")[0].strip()
                # CMSコメント除外
                lines = [l.strip() for l in after_q.split("\n")
                         if l.strip() and not re.match(r'^「.+」$', l.strip())
                         and not l.strip().startswith("ここから")
                         and not l.strip().startswith("ここまで")]
                answer = "\n".join(lines).strip()

    result["answer"] = answer

    # ---- 問い合わせ先: dl > dd ----
    inquiry = ""
    if contents_in:
        for dl in contents_in.find_all("dl"):
            if "この記事に関するお問い合わせ先" in dl.get_text():
                dd = dl.find("dd")
                if dd:
                    inquiry = dd.get_text(separator="\n", strip=True)
                    inquiry = inquiry.replace("お問い合わせフォーム", "").strip()
                break

    result["inquiry"] = inquiry
    return result


def save_to_excel(faq_data, output_path):
    wb = openpyxl.Workbook()
    hf = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="2F5496")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    # サマリー
    ws0 = wb.active
    ws0.title = "サマリー"
    n_with = sum(1 for d in faq_data if d.get("answer"))
    ws0.cell(row=1, column=1, value="京丹後市FAQ収集結果").font = Font(name="Arial", size=12, bold=True)
    ws0.cell(row=2, column=1, value="合計FAQ数").font = cf
    ws0.cell(row=2, column=2, value=len(faq_data)).font = cf
    ws0.cell(row=3, column=1, value="回答取得済み").font = cf
    ws0.cell(row=3, column=2, value=n_with).font = cf
    ri = 5
    ws0.cell(row=ri, column=1, value="カテゴリ別件数").font = Font(name="Arial", size=11, bold=True)
    cat_counts = defaultdict(int)
    for d in faq_data:
        cat_counts[f"{d['main_category']} > {d['sub_category']}"] += 1
    for cat, cnt in sorted(cat_counts.items()):
        ri += 1
        ws0.cell(row=ri, column=1, value=cat).font = cf
        ws0.cell(row=ri, column=2, value=cnt).font = cf
    ws0.column_dimensions["A"].width = 40
    ws0.column_dimensions["B"].width = 10

    # FAQ一覧
    ws1 = wb.create_sheet("FAQ一覧")
    headers = ["No.", "大カテゴリ", "サブカテゴリ", "質問", "回答", "問い合わせ先", "URL"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb
    for ri, d in enumerate(faq_data, 2):
        vals = [ri-1, d.get("main_category",""), d.get("sub_category",""),
                d.get("question",""), d.get("answer","")[:2000],
                d.get("inquiry",""), d.get("url","")]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.font, c.alignment, c.border = cf, ca, tb
    for col, w in [("A",6),("B",18),("C",18),("D",55),("E",80),("F",40),("G",50)]:
        ws1.column_dimensions[col].width = w
    ws1.auto_filter.ref = f"A1:G{len(faq_data)+1}"
    ws1.freeze_panes = "A2"
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        ml = max((str(c.value).count('\n')+1 if c.value and c.column in (5,6) else 1) for c in row)
        ws1.row_dimensions[row[0].row].height = min(max(ml*15,30),200)
    wb.save(output_path)
    print(f"\n出力: {output_path}")


def main():
    print("=" * 60)
    print("京丹後市FAQ収集スクリプト v5")
    print("=" * 60)

    print("\n【STEP 1】カテゴリ取得")
    categories = get_category_pages(FAQ_TOP_URL)
    if not categories:
        return

    print("\n【STEP 2】FAQリンク取得")
    all_links = []
    for cat in categories:
        time.sleep(REQUEST_DELAY)
        links = get_faq_links(cat["url"])
        print(f"  [{cat['main_category']}] {cat['sub_category']}: {len(links)}件")
        for lk in links:
            lk["main_category"] = cat["main_category"]
            lk["sub_category"] = cat["sub_category"]
        all_links.extend(links)
    seen = set()
    unique = [lk for lk in all_links if lk["url"] not in seen and not seen.add(lk["url"])]
    print(f"  合計: {len(unique)}件")

    print("\n【STEP 3】FAQ詳細取得")
    faq_data, empty = [], 0
    for i, lk in enumerate(unique, 1):
        time.sleep(REQUEST_DELAY)
        if i % 10 == 0 or i == len(unique):
            print(f"  {i}/{len(unique)}")
        detail = extract_faq_detail(lk["url"])
        if detail:
            detail["main_category"] = lk["main_category"]
            detail["sub_category"] = lk["sub_category"]
            if not detail.get("question"):
                detail["question"] = lk["title"]
            if not detail.get("answer"):
                empty += 1
                print(f"    [WARN] 回答空: {lk['title'][:40]}")
            faq_data.append(detail)
    print(f"\n  結果: {len(faq_data)}件（回答あり: {len(faq_data)-empty}, 空: {empty}）")

    print("\n【STEP 4】Excel出力")
    save_to_excel(faq_data, OUTPUT_FILE)
    print(f"\n完了！ {len(faq_data)}件 → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()