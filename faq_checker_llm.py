"""
自治体FAQ正誤チェックツール v2（LLM直接判定＋2次検証）
=======================================================
自治体FAQの各回答に対して、Gemini APIが自身の知識と
総務省FAQの情報を組み合わせて正誤を判定する。
1次チェック → 2次検証（セカンドオピニオン）の二段構えで精度を担保。

使い方:
  1. .envファイルにGEMINI_API_KEYを設定
  2. 下記「ユーザー設定」セクションでファイルパス・モデル等を変更
  3. python faq_checker_llm.py
"""

import os
import re
import json
import time
import threading
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# ★★★ ユーザー設定（ここを変更してください）★★★
# ============================================================

# Geminiモデル名

GEMINI_MODEL = "gemini-3.1-pro-preview"

# 自治体名（出力ファイル名やログに使用）
CITY_NAME = "京丹後市"

# 入力ファイルパス
CITY_FAQ_PATH = "京丹後市FAQ.xlsx"
SOUMU_FAQ_PATH = "総務省FAQ.xlsx"

# 出力ファイルパス
OUTPUT_PATH = "FAQ正誤チェック結果_京丹後市_gemini3.1prov2.xlsx"

# 並列実行数（APIのレート制限に応じて調整。2〜5推奨）
MAX_WORKERS = 4

# LLM生成設定
TEMPERATURE = 0.1
MAX_OUTPUT_TOKENS = 2048

# ============================================================
# 以下は通常変更不要
# ============================================================

load_dotenv()

_print_lock = threading.Lock()
def safe_print(msg):
    with _print_lock:
        print(msg)


# ============================================================
# 最新の制度情報（プロンプトに埋め込む参照データ）
# ここを更新することで、法改正への対応を維持できる
# ============================================================
SYSTEM_KNOWLEDGE = """
■ 主要な法改正・制度変更（チェック時に必ず参照すること）

【届出・手続き関連】
・戸籍届出の押印任意化: 令和3年9月1日施行。出生届・死亡届・婚姻届・離婚届・転籍届など全ての戸籍届出で押印は任意。「印鑑をお持ちください」「届出人の印鑑」等を必須として案内している場合 → 誤りあり
・戸籍謄本の添付不要化: 令和6年3月1日施行。本籍地以外での届出時も戸籍謄本の添付は原則不要。「戸籍謄本が必要」と案内している場合 → 誤りあり
・成年年齢の引き下げ: 令和4年4月1日施行。成年年齢は18歳（旧: 20歳）。婚姻届の証人等で「20歳以上」としている場合 → 誤りあり
・行政手続きの押印廃止: 令和3年以降、国の方針で多くの行政手続きの押印が廃止。運転免許の自主返納等でも印鑑は原則不要。

【年金・保険関連】
・年金手帳の廃止: 令和4年4月1日。「基礎年金番号通知書」に変更。「年金手帳」を必要書類として案内 → 誤りあり
・健康保険証の新規発行廃止: 令和6年12月2日。マイナ保険証または資格確認書に移行。「保険証をお送りします」等の案内 → 要確認
・20歳到達時の国民年金加入: 令和元年10月以降、届出不要（自動加入）。「届出が必要」と案内 → 誤りあり

【制度名称の変更】
・「障害程度区分」→「障害支援区分」（平成26年4月改正）
・「介護療養型医療施設」→「介護医療院」（令和6年3月末で完全廃止）
・「社団法人」→「一般社団法人」（公益法人制度改革による移行）

【手当・給付金の金額（令和6年度）】
・特別障害者手当: 月額28,840円
・障害児福祉手当: 月額15,690円
・特別児童扶養手当（1級）: 月額55,350円
・特別児童扶養手当（2級）: 月額36,860円
・出産育児一時金: 50万円（令和5年4月〜）
・児童手当: 第1子・第2子 月額10,000円（3歳未満15,000円）、第3子以降30,000円（令和6年10月〜）
※ 上記と異なる金額が記載されている場合 → 誤りあり

【介護保険関連】
・介護保険料の負担割合: 通常1割（一定以上所得者は2割または3割）
・2年以上滞納時: 利用者負担が3割または4割に引き上げ（令和5年介護保険法改正で4割の場合あり）
・「1割から3割になる」のみの説明は不完全（現行制度では通常時でも最大3割のため）→ 要確認

【税制関連】
・2025年（令和7年）税制改正: 基礎控除が最大95万円に引き上げ（所得132万円以下の場合）
・給与所得控除の最低保障額: 55万円→65万円に引き上げ
・所得税の非課税ライン: 給与収入160万円（基礎控除95万円＋給与所得控除65万円）
・扶養控除の所得要件: 48万円→58万円に引き上げ（給与収入103万円→123万円）
・住民税の基礎控除: 据え置き（所得税と異なる）
※ 自治体の数値が従来の103万円等と異なる場合、税制改正を反映している可能性あり → 要確認（誤りと断定しない）

【国民健康保険関連】
・国保税の特別徴収（年金天引き）から口座振替への変更: 多くの自治体で申し出により可能。これは誤りではない。
・加入・脱退の届出期限: 14日以内（総務省FAQ明記）
"""


# ============================================================
# Gemini API
# ============================================================
def make_generation_url(model, api_key):
    return f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"


def call_gemini(prompt, generation_url, temperature=0.1, max_tokens=2048):
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": temperature, "maxOutputTokens": max_tokens},
        "tools": [{"google_search": {}}],
    }
    for attempt in range(5):
        try:
            resp = requests.post(generation_url, json=payload, timeout=120)
            if resp.status_code == 429:
                wait = (attempt + 1) * 15
                safe_print(f"      レート制限。{wait}秒待機...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()
            # Google Search Grounding使用時、partsに複数のブロックが返る場合がある
            # textタイプのパートからテキストを抽出
            parts = data["candidates"][0]["content"]["parts"]
            text_parts = [p["text"] for p in parts if "text" in p]
            text = "\n".join(text_parts).strip()
            if text.startswith("```"):
                text = re.sub(r"^```(?:json)?\s*", "", text)
                text = re.sub(r"\s*```$", "", text)
            return text
        except requests.exceptions.HTTPError:
            if resp.status_code == 429:
                time.sleep((attempt + 1) * 15)
                continue
            if attempt == 4: raise
            time.sleep(5)
        except Exception:
            if attempt == 4: raise
            time.sleep(5)
    raise RuntimeError("5回リトライ後も失敗")


# ============================================================
# データ読み込み
# ============================================================
def load_faq(path):
    all_data = []
    xls = pd.ExcelFile(path)
    for sheet in xls.sheet_names:
        if sheet in ("サマリー",):
            continue
        df = pd.read_excel(path, sheet_name=sheet)
        if "質問" in df.columns and "回答" in df.columns:
            df["シート"] = sheet
            all_data.append(df)
    if not all_data:
        raise ValueError(f"「質問」「回答」列を含むシートが見つかりません: {path}")
    result = pd.concat(all_data, ignore_index=True)
    if "サブカテゴリ" not in result.columns:
        result["サブカテゴリ"] = result.get("大カテゴリ", pd.Series(["その他"] * len(result)))
    return result


def build_soumu_summary(soumu_df):
    summaries = {}
    for sheet in soumu_df["シート"].unique():
        subset = soumu_df[soumu_df["シート"] == sheet]
        lines = []
        for _, row in subset.iterrows():
            subcat = row.get("サブカテゴリ", "")
            q = str(row["質問"])[:100]
            a = str(row["回答"])[:300]
            lines.append(f"[{subcat}] Q: {q}\nA: {a}")
        summaries[sheet] = "\n\n".join(lines)
    return summaries


def find_relevant_soumu_categories(city_subcat, city_question, soumu_summaries):
    keyword_map = {
        "国民健康保険": ["国民健康保険 FAQ"], "国保": ["国民健康保険 FAQ"],
        "健康保険": ["国民健康保険 FAQ"], "高額療養": ["国民健康保険 FAQ"],
        "年金": ["年金 FAQ"], "国民年金": ["年金 FAQ"],
        "住民税": ["税金 FAQ"], "市県民税": ["税金 FAQ"], "市民税": ["税金 FAQ"],
        "固定資産税": ["税金 FAQ"], "軽自動車税": ["税金 FAQ"],
        "税": ["税金 FAQ"], "所得税": ["税金 FAQ"],
        "福祉": ["福祉 FAQ"], "介護": ["福祉 FAQ"], "障害": ["福祉 FAQ"], "手当": ["福祉 FAQ"],
        "戸籍": ["届出・証明 FAQ"], "届出": ["届出・証明 FAQ"],
        "婚姻": ["届出・証明 FAQ"], "出生": ["届出・証明 FAQ"],
        "死亡届": ["届出・証明 FAQ"], "離婚": ["届出・証明 FAQ"],
        "住民票": ["届出・証明 FAQ"], "転出": ["届出・証明 FAQ"], "転入": ["届出・証明 FAQ"],
        "印鑑": ["届出・証明 FAQ"],
        "マイナンバー": ["マイナンバー FAQ"], "防災": ["防災・安全 FAQ"],
        "子育て": ["子育て・教育 FAQ"], "児童手当": ["子育て・教育 FAQ"],
        "保育": ["子育て・教育 FAQ"],
        "出産": ["子育て・教育 FAQ", "国民健康保険 FAQ"],
        "給付金": ["給付金 FAQ"], "登記": ["住宅・土地 FAQ"],
        "相続": ["住宅・土地 FAQ"], "住宅": ["住宅・土地 FAQ"],
    }
    matched = set()
    search_text = f"{city_subcat} {city_question}"
    for keyword, categories in keyword_map.items():
        if keyword in search_text:
            matched.update(categories)
    return [cat for cat in matched if cat in soumu_summaries]


# ============================================================
# 1次チェック
# ============================================================
def check_one_faq(task):
    idx, row, soumu_context, total, city_name, generation_url = task
    question = str(row["質問"])
    answer = str(row["回答"])
    category = str(row.get("サブカテゴリ", ""))
    ref_text = soumu_context[:3000] if soumu_context else "（関連する総務省FAQなし）"

    prompt = f"""あなたは日本の行政制度に精通した専門家です。
以下の自治体FAQの回答が正確かどうかを、制度情報リファレンスと総務省の参照FAQを使って検証してください。

{SYSTEM_KNOWLEDGE}

■ 判定カテゴリと基準

【誤りあり】以下のいずれかに該当する場合：
- 上記の制度情報リファレンスに照らして、廃止された義務を必須として案内している
- 法改正で変更された数値・要件が旧制度のまま記載されている
- 廃止された制度名をそのまま使用している
- 明らかに事実と異なる説明がある
※ 該当する場合は必ず「誤りあり」とすること

【要確認】以下のいずれかに該当する場合：
- 制度上の誤解を招く表現がある場合
- 古い情報が残っている可能性があるが断定できない場合
- 総務省FAQや法令に記載されている重要な条件・例外・期限が記載されていない場合（例: 届出期限「14日以内」、年齢要件、所得要件、対象者の除外条件など）
- 総務省FAQでは説明されている制度の重要な側面（対象者の範囲、除外条件、負担割合の区分など）が省略されている場合
- 用語や表現が厳密には不正確な場合（例: 法令上の正式な制度名と異なる用語の使用、「所在地」と「住所地」の混同など）

【問題なし】以下の場合：
- 回答内容に誤りがなく、重要な条件・例外の欠落もない場合

【対象外】自治体固有の情報（窓口案内、電話番号、独自サービス等）のみの内容

---
【チェック対象：{city_name}FAQ】
カテゴリ: {category}
質問: {question}
回答: {answer}

---
【総務省 参照FAQ】
{ref_text}

---
以下のJSON形式のみで回答してください。JSON以外は一切出力しないでください。
{{
  "判定": "対象外" or "問題なし" or "要確認" or "誤りあり",
  "指摘事項": "具体的な指摘内容。問題なし・対象外の場合は空文字",
  "根拠": "判定の根拠となる法令・制度情報。問題なし・対象外の場合は空文字",
  "該当箇所": "問題がある場合、自治体回答の該当部分を引用。問題なし・対象外の場合は空文字"
}}"""

    try:
        text = call_gemini(prompt, generation_url)
        result = json.loads(text)
    except json.JSONDecodeError:
        result = _try_extract_partial_json(text)
    except Exception as e:
        result = {"判定": "APIエラー", "指摘事項": str(e), "根拠": "", "該当箇所": ""}

    judgment = result.get("判定", "不明")
    safe_print(f"    [{idx+1}/{total}] {judgment} - {question[:45]}...")
    return {
        "idx": idx, "カテゴリ": category, "質問": question, "回答": answer,
        "判定": judgment, "指摘事項": result.get("指摘事項", ""),
        "根拠": result.get("根拠", ""), "該当箇所": result.get("該当箇所", ""),
    }


# ============================================================
# 2次検証
# ============================================================
def verify_one_faq(task):
    idx, row_data, generation_url = task
    question = row_data["質問"]
    answer = row_data["回答"]
    category = row_data["カテゴリ"]
    first_judgment = row_data["判定"]
    first_reason = row_data["指摘事項"]
    first_basis = row_data["根拠"]

    prompt = f"""あなたは行政制度の専門家であり、FAQチェック結果のレビュアーです。

1次チェックで「{first_judgment}」と判定されたFAQを再検証してください。

{SYSTEM_KNOWLEDGE}

■ 再検証ルール

【判定を維持すべき場合（格下げ禁止）】
- 1次チェックの指摘が上記の制度情報リファレンスに記載されている法改正に該当する場合
  例: 押印の任意化、戸籍謄本の添付不要化、成年年齢の引き下げ、手当金額の変更、制度名称の変更
  → これらは事実として確定しており、「問題なし」への格下げは禁止

【「問題なし」に変更してよい場合】
- 1次チェックの指摘自体に事実誤認がある場合（法令の解釈が間違っている等）
- 自治体独自の運用として許容範囲であることが確認できた場合

【「問題なし」に変更してはいけない場合】
- 1次チェックが「重要な条件・例外・期限の記載漏れ」を指摘している場合 → 要確認を維持
- 1次チェックが「用語の不正確さ」を指摘している場合 → 要確認を維持
- 1次チェックの指摘が制度情報リファレンスに記載されている法改正に該当する場合 → 誤りありを維持

---
【自治体FAQ】
カテゴリ: {category}
質問: {question}
回答: {answer}

【1次チェック結果】
判定: {first_judgment}
指摘事項: {first_reason}
根拠: {first_basis}

---
以下のJSON形式のみで回答してください。
{{
  "最終判定": "{first_judgment}" or "問題なし",
  "判定変更理由": "維持の場合は維持理由を、変更の場合は変更理由を記述",
  "修正後の指摘事項": "維持の場合は1次の指摘を精緻化。問題なしに変更した場合は空文字",
  "修正後の根拠": "維持の場合は1次の根拠を精緻化。問題なしに変更した場合は空文字"
}}"""

    try:
        text = call_gemini(prompt, generation_url)
        result = json.loads(text)
    except json.JSONDecodeError:
        final = first_judgment
        for label in ["問題なし", "誤りあり", "要確認"]:
            if f'"{label}"' in text:
                final = label
                break
        result = {
            "最終判定": final,
            "判定変更理由": f"JSON解析失敗（推定）: {text[:200]}",
            "修正後の指摘事項": first_reason if final != "問題なし" else "",
            "修正後の根拠": first_basis if final != "問題なし" else "",
        }
    except Exception as e:
        result = {
            "最終判定": first_judgment,
            "判定変更理由": f"API失敗（1次判定を維持）: {str(e)}",
            "修正後の指摘事項": first_reason,
            "修正後の根拠": first_basis,
        }

    final_judgment = result.get("最終判定", first_judgment)
    changed = "→変更" if final_judgment != first_judgment else "→維持"
    safe_print(f"    [再検証] {first_judgment}{changed}:{final_judgment} - {question[:40]}...")
    return {
        "idx": idx,
        "最終判定": final_judgment,
        "判定変更理由": result.get("判定変更理由", ""),
        "修正後の指摘事項": result.get("修正後の指摘事項", ""),
        "修正後の根拠": result.get("修正後の根拠", ""),
    }


def run_verification(df, generation_url, max_workers):
    targets = df[df["判定"].isin(["誤りあり", "要確認"])].copy()
    if len(targets) == 0:
        print("  2次検証の対象なし")
        return df

    print(f"  {len(targets)}件を2次検証します...")
    tasks = [(idx, row, generation_url) for idx, row in targets.iterrows()]
    verify_results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(verify_one_faq, task): task[0] for task in tasks}
        for future in as_completed(futures):
            res = future.result()
            verify_results[res["idx"]] = res

    changed_count = 0
    for idx, vr in verify_results.items():
        original = df.loc[idx, "判定"]
        final = vr["最終判定"]
        if final != original:
            changed_count += 1
        df.loc[idx, "判定"] = final
        df.loc[idx, "指摘事項"] = vr["修正後の指摘事項"] if final != "問題なし" else ""
        df.loc[idx, "根拠"] = vr["修正後の根拠"] if final != "問題なし" else ""
        df.loc[idx, "該当箇所"] = df.loc[idx, "該当箇所"] if final != "問題なし" else ""

    print(f"  2次検証完了: {changed_count}件の判定が変更されました")
    return df


# ============================================================
# ユーティリティ
# ============================================================
def _try_extract_partial_json(text):
    judgment = "解析エラー"
    for label in ["誤りあり", "要確認", "問題なし", "対象外"]:
        if f'"{label}"' in text:
            judgment = label
            break
    reason = ""
    m = re.search(r'"指摘事項"\s*:\s*"([^"]*)', text)
    if m: reason = m.group(1)
    basis = ""
    m = re.search(r'"根拠"\s*:\s*"([^"]*)', text)
    if m: basis = m.group(1)
    return {"判定": judgment, "指摘事項": reason or f"JSON解析失敗: {text[:200]}", "根拠": basis, "該当箇所": ""}


# ============================================================
# Excel出力
# ============================================================
def export_excel(df, output_path):
    sort_order = {"誤りあり": 0, "要確認": 1, "APIエラー": 2, "解析エラー": 3, "問題なし": 4, "対象外": 5}
    df["sort_key"] = df["判定"].map(sort_order).fillna(6)
    df = df.sort_values("sort_key").drop(columns=["sort_key", "idx"])
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "要対応"
    actionable = df[df["判定"].isin(["誤りあり", "要確認"])].copy()
    _write_sheet(ws1, actionable, ["カテゴリ", "判定", "質問", "回答", "指摘事項", "根拠"])
    ws2 = wb.create_sheet("全件結果")
    _write_sheet(ws2, df, ["カテゴリ", "判定", "質問", "回答", "指摘事項", "根拠", "該当箇所"])
    ws3 = wb.create_sheet("サマリー")
    _write_summary(ws3, df)
    wb.save(output_path)
    err = len(actionable[actionable["判定"] == "誤りあり"])
    warn = len(actionable[actionable["判定"] == "要確認"])
    print(f"  出力先: {output_path}")
    print(f"  要対応: {len(actionable)}件（誤りあり: {err}件 / 要確認: {warn}件）")


def _write_sheet(ws, df, columns):
    hf = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bf = Font(name="Arial", size=10)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    fm = {"誤りあり": PatternFill("solid", fgColor="FFC7CE"), "要確認": PatternFill("solid", fgColor="FFEB9C"),
          "問題なし": PatternFill("solid", fgColor="C6EFCE"), "対象外": PatternFill("solid", fgColor="D9D9D9")}
    cw = {"カテゴリ": 18, "判定": 14, "質問": 40, "回答": 55, "指摘事項": 65, "根拠": 75, "該当箇所": 40}
    for ci, h in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=h); c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = tb
    for ri, (_, rd) in enumerate(df.iterrows(), 2):
        for ci, cn in enumerate(columns, 1):
            c = ws.cell(row=ri, column=ci, value=rd.get(cn, ""))
            c.font = bf; c.alignment = Alignment(vertical="top", wrap_text=True); c.border = tb
        f = fm.get(str(rd.get("判定", "")))
        if f:
            for ci in range(1, len(columns) + 1): ws.cell(row=ri, column=ci).fill = f
        ml = max(len(str(rd.get(c, ""))) for c in columns)
        ws.row_dimensions[ri].height = max(min(ml / 3, 200), 50)
    for ci, h in enumerate(columns, 1):
        cl = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[cl].width = cw.get(h, 15)
    lc = chr(64 + len(columns))
    ws.auto_filter.ref = f"A1:{lc}{len(df)+1}"; ws.freeze_panes = "A2"


def _write_summary(ws, df):
    hf = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bf = Font(name="Arial", size=10); blf = Font(bold=True, name="Arial", size=10)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    hs = ["カテゴリ", "分析件数", "誤りあり", "要確認", "問題なし", "対象外"]
    for ci, h in enumerate(hs, 1):
        c = ws.cell(row=1, column=ci, value=h); c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = tb
    cats = df["カテゴリ"].unique()
    for ri, cat in enumerate(cats, 2):
        s = df[df["カテゴリ"] == cat]
        d = [cat, len(s), len(s[s["判定"]=="誤りあり"]), len(s[s["判定"]=="要確認"]),
             len(s[s["判定"]=="問題なし"]), len(s[s["判定"]=="対象外"])]
        for ci, v in enumerate(d, 1):
            c = ws.cell(row=ri, column=ci, value=v); c.font = bf
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left"); c.border = tb
    tr = len(cats) + 2
    ws.cell(row=tr, column=1, value="合計").font = blf; ws.cell(row=tr, column=1).border = tb
    for ci in range(2, 7):
        cl = ws.cell(row=tr, column=ci).column_letter
        c = ws.cell(row=tr, column=ci); c.value = f"=SUM({cl}2:{cl}{tr-1})"
        c.font = blf; c.alignment = Alignment(horizontal="center"); c.border = tb
    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F"]: ws.column_dimensions[col].width = 12


# ============================================================
# メイン処理
# ============================================================
def main():
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError(".envファイルにGEMINI_API_KEYが設定されていません")

    generation_url = make_generation_url(GEMINI_MODEL, api_key)

    start_time = time.time()
    print("=" * 60)
    print(f"  {CITY_NAME} FAQ正誤チェックツール v2")
    print(f"  モデル: {GEMINI_MODEL}")
    print(f"  並列数: {MAX_WORKERS}")
    print("=" * 60)

    # ステップ1: データ読み込み
    print("\nステップ1: データ読み込み")
    city_df = load_faq(CITY_FAQ_PATH)
    soumu_df = load_faq(SOUMU_FAQ_PATH)
    print(f"  {CITY_NAME}FAQ: {len(city_df)}件")
    print(f"  総務省FAQ: {len(soumu_df)}件")

    # ステップ2: 総務省FAQ参照データ構築
    print("\nステップ2: 総務省FAQ参照データ構築")
    soumu_summaries = build_soumu_summary(soumu_df)
    print(f"  {len(soumu_summaries)}カテゴリの参照データを構築")

    # ステップ3: 1次チェック
    print(f"\nステップ3: LLMによる1次チェック（{MAX_WORKERS}並列）")
    print(f"  {len(city_df)}件のFAQをチェックします...")

    tasks = []
    for i, row in city_df.iterrows():
        subcat = str(row.get("サブカテゴリ", ""))
        question = str(row["質問"])
        relevant_cats = find_relevant_soumu_categories(subcat, question, soumu_summaries)
        soumu_context = "\n\n---\n\n".join(f"【{cat}】\n{soumu_summaries[cat]}" for cat in relevant_cats)
        tasks.append((i, row, soumu_context, len(city_df), CITY_NAME, generation_url))

    results = [None] * len(city_df)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(check_one_faq, task): task[0] for task in tasks}
        for future in as_completed(futures):
            res = future.result()
            results[res["idx"]] = res

    df = pd.DataFrame(results)
    print(f"\n  === 1次チェック結果 ===")
    for label in ["誤りあり", "要確認", "問題なし", "対象外", "APIエラー", "解析エラー"]:
        count = len(df[df["判定"] == label])
        if count > 0: print(f"  {label}: {count}件")

    # ステップ4: 2次検証
    print(f"\nステップ4: 2次検証（セカンドオピニオン）")
    df = run_verification(df, generation_url, MAX_WORKERS)
    print(f"\n  === 最終判定結果 ===")
    for label in ["誤りあり", "要確認", "問題なし", "対象外", "APIエラー", "解析エラー"]:
        count = len(df[df["判定"] == label])
        if count > 0: print(f"  {label}: {count}件")

    # ステップ5: Excel出力
    print(f"\nステップ5: Excel出力")
    export_excel(df, OUTPUT_PATH)

    elapsed = time.time() - start_time
    print(f"\n  処理時間: {elapsed:.0f}秒（{elapsed/60:.1f}分）")
    print("=" * 60)
    print("  処理完了")
    print("=" * 60)


if __name__ == "__main__":
    main()
